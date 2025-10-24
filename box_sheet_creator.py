from dataclasses import fields

from constraints.field_name_mapper import BOX_CODE_FIELD_NAME, CABLE_SKIP_COUNT_FIELD_NAME, CABLE_PORT_START_FIELD_NAME, \
    CABLE_PORT_END_FIELD_NAME, CABLE_CODE_FIELD_NAME, CABLE_TYPE_FIELD_NAME, CABLE_SECTION_FIELD_NAME, \
    BOX_IN_START_FIELD_NAME, BOX_IN_END_FIELD_NAME, BOX_TYPE_FIELD_NAME, CABLE_LEVEL_FIELD_NAME, \
    CABLE_EXTREMITY_FIELD_NAME
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from data_service import data_service_cable, data_service_box
from utils import excel_utils

def create_box_sheet(wb, topo_box_col, box_data):
    if topo_box_col == 'A':
        return _create_sro_sheet(wb, box_data)
    return None

def _create_sro_sheet(wb, sro_data):
    sheet_name = sro_data[BOX_CODE_FIELD_NAME]
    # 获取工作簿对象（从当前工作表反向获取）

    # 检查子sheet是否已存在，不存在则创建
    if sheet_name not in wb.sheetnames:
        ws_sro = wb.create_sheet(title=sheet_name)
        ws_sro.sheet_view.zoomScale = SHEET_ZOOM_SCALE
        ws_sro.append([])
        start_col = 1
        # 绘制返回按钮
        create_return_topo_cell(ws_sro)
        # 绘制表头
        data_1st_row = draw_sro_table_header(ws_sro)
        fill_d1_data(ws_sro, data_1st_row, start_col)
        fill_d2_data(ws_sro, data_1st_row, start_col + COL_LOOP + 5)
        fill_d3_data(ws_sro, data_1st_row, start_col + COL_LOOP * 2)
        ws_sro.freeze_panes = "A4"
        return sheet_name
    return None


def fill_d2_data(ws_sro, data_1st_row, start_col):
    _1st_segments_on_d2_section = data_service_cable.get_all_1st_segments_on_d2_section_order_by_skip_count_asc()
    if _1st_segments_on_d2_section is not None and not _1st_segments_on_d2_section.empty:
        for _idx, _1st_segment in _1st_segments_on_d2_section.iterrows():
            skip_count = int(_1st_segment[CABLE_SKIP_COUNT_FIELD_NAME])
            section = _1st_segment[CABLE_SECTION_FIELD_NAME]
            port_start = int(_1st_segment[CABLE_PORT_START_FIELD_NAME])
            port_end = int(_1st_segment[CABLE_PORT_END_FIELD_NAME])
            fiber_to_fill_amt = port_end - port_start + 1
            for i in range(0, fiber_to_fill_amt):
                no = i + 1
                row_no = data_1st_row + skip_count + i
                section_cell = ws_sro.cell(row=row_no, column=start_col)
                section_cell.value = section

                type_cell = ws_sro.cell(row=row_no, column=start_col + 1)
                type_cell.value = _1st_segment[CABLE_TYPE_FIELD_NAME]

                no_cell = ws_sro.cell(row=row_no, column=start_col + 2)
                no_cell.value = no

                t = int(i / 12) + 1
                t_cell = ws_sro.cell(row=row_no, column=start_col + 3)
                t_cell.value = t
                t_cell.fill = PatternFill(fill_type="solid", start_color=get_fiber_color(t - 1))

                f = i % 12 + 1
                f_cell = ws_sro.cell(row=row_no, column=start_col + 4)
                f_cell.value = f
                f_cell.fill = PatternFill(fill_type="solid", start_color=get_fiber_color(f - 1))
            boxs_on_section = data_service_box.get_all_boxs_on_section_by_orders(section=section,
                                                                                 sort_by=[BOX_IN_START_FIELD_NAME])
            if boxs_on_section is not None and not boxs_on_section.empty:
                for _box_idx, _box in boxs_on_section.iterrows():
                    fill_closure_port_on_section(ws_sro, data_1st_row, start_col + 5, _box)


def fill_d1_data(ws_sro, data_1st_row, start_col):
    _1st_segments = data_service_cable.get_all_1st_segments_on_d1_section_order_by_skip_count_asc()
    # 初始化数据
    if _1st_segments is not None and not _1st_segments.empty:
        sro_port_idx = 1
        odf_code_idx = 1
        odf_port_idx = 1
        for _idx, _1st_segment in _1st_segments.iterrows():
            skip_count = int(_1st_segment[CABLE_SKIP_COUNT_FIELD_NAME])
            section = _1st_segment[CABLE_SECTION_FIELD_NAME]
            port_start = int(_1st_segment[CABLE_PORT_START_FIELD_NAME])
            port_end = int(_1st_segment[CABLE_PORT_END_FIELD_NAME])

            fiber_to_fill_amt = port_end - port_start + 1
            for i in range(0, fiber_to_fill_amt):
                row_no = data_1st_row + skip_count + i

                sro_port_cell = ws_sro.cell(row=row_no, column=start_col)
                sro_port_cell.value = sro_port_idx

                odf_code_cell = ws_sro.cell(row=row_no, column=start_col + 1)
                odf_code_cell.value = f"ODF{'%02d' % odf_code_idx}"

                odf_port_cell = ws_sro.cell(row=row_no, column=start_col + 2)
                odf_port_cell.value = odf_port_idx

                splice_cell = ws_sro.cell(row=row_no, column=start_col + 3)
                splice_cell.value = 'S'

                cable_code_cell = ws_sro.cell(row=row_no, column=start_col + 5)
                cable_code_cell.value = section

                cable_type_cell = ws_sro.cell(row=row_no, column=start_col + 6)
                cable_type_cell.value = _1st_segment[CABLE_TYPE_FIELD_NAME]

                cable_no_cell = ws_sro.cell(row=row_no, column=start_col + 7)
                cable_no_cell.value = i + 1

                cable_t_cell = ws_sro.cell(row=row_no, column=start_col + 8)
                t = int(i / 12) + 1
                cable_t_cell.value = t
                cable_t_cell.fill = PatternFill(fill_type="solid", start_color=get_fiber_color(t - 1))

                cable_f_cell = ws_sro.cell(row=row_no, column=start_col + 9)
                f = i % 12 + 1
                cable_f_cell.value = f
                cable_f_cell.fill = PatternFill(fill_type="solid", start_color=get_fiber_color(f - 1))

                sro_port_idx += 1
                odf_code_idx = odf_code_idx if odf_port_idx < ODF_MAX_PORT_NO else odf_code_idx + 1
                odf_port_idx = odf_port_idx + 1 if odf_port_idx < ODF_MAX_PORT_NO else 1

            boxs_on_section = data_service_box.get_all_boxs_on_section_by_orders(section=section,
                                                                                 sort_by=[BOX_IN_START_FIELD_NAME])
            if boxs_on_section is not None and not boxs_on_section.empty:
                for _box_idx, _box in boxs_on_section.iterrows():
                    fill_closure_port_on_section(ws_sro, data_1st_row, start_col + COL_LOOP, _box)

    # fill_d2_fiber_data(ws_sro, data_1st_row, start_col + 15)
    # fill_d3_fiber_data(ws_sro, data_1st_row, start_col + 25)


def fill_closure_port_on_section(ws_sro, data_1st_row, start_col, box_data):
    in_start = box_data[BOX_IN_START_FIELD_NAME]
    in_end = box_data[BOX_IN_END_FIELD_NAME]
    skip_count = int(box_data[CABLE_SKIP_COUNT_FIELD_NAME])
    port_to_fill_amt = in_end - in_start + 1
    for i in range(0, port_to_fill_amt):
        no = i + 1
        row_no = data_1st_row + skip_count + i
        box_port_cell = ws_sro.cell(row=row_no, column=start_col)
        box_port_cell.value = box_data[BOX_CODE_FIELD_NAME]

        box_type_cell = ws_sro.cell(row=row_no, column=start_col + 1)
        box_type_cell.value = box_data[BOX_TYPE_FIELD_NAME]

        box_no_cell = ws_sro.cell(row=row_no, column=start_col + 2)
        box_no_cell.value = no

        splice_cell = ws_sro.cell(row=row_no, column=start_col + 3)
        splice_cell.value = 'S'


def fill_d2_fiber_data(ws_sro, data_1st_row, start_col):
    d2_cables = data_service_cable.get_all_d2_cables_order_by_skip_count()
    if d2_cables is not None and not d2_cables.empty:
        for _d2_idx, d2_cable in d2_cables.iterrows():
            section = d2_cable[CABLE_SECTION_FIELD_NAME]
            _type = d2_cable[CABLE_TYPE_FIELD_NAME]
            skip_count = int(d2_cable[CABLE_SKIP_COUNT_FIELD_NAME])
            port_start = int(d2_cable[CABLE_PORT_START_FIELD_NAME])
            port_end = int(d2_cable[CABLE_PORT_END_FIELD_NAME])
            for i in range(skip_count, skip_count + port_end - port_start + 1):
                code_cell = ws_sro.cell(row=data_1st_row + i, column=start_col)
                code_cell.value = section

                type_cell = ws_sro.cell(row=data_1st_row + i, column=start_col + 1)
                type_cell.value = _type

                on_cell = ws_sro.cell(row=data_1st_row + i, column=start_col + 2)
                on_cell.value = i - skip_count + 1

                t = int((i - skip_count + 1 - 0.01) / 12) + 1
                t_cell = ws_sro.cell(row=data_1st_row + i, column=start_col + 3)
                t_cell.value = t
                t_cell.fill = PatternFill(fill_type="solid", start_color=get_fiber_color(t))

                f = (i - skip_count) % 12 + 1
                f_cell = ws_sro.cell(row=data_1st_row + i, column=start_col + 4)
                f_cell.value = f
                f_cell.fill = PatternFill(fill_type="solid", start_color=get_fiber_color(f))

            boxs_on_section = data_service_box.get_all_boxs_on_section_by_orders(section=section,
                                                                                 sort_by=[BOX_IN_START_FIELD_NAME])
            if boxs_on_section is not None and not boxs_on_section.empty:
                for _box_idx, _box in boxs_on_section.iterrows():
                    fill_closure_port_on_section(ws_sro, data_1st_row, start_col + 5, _box)


def fill_d3_data(ws_sro, data_1st_row, start_col):
    _1st_segments_on_d3_cables = data_service_cable.get_all_1st_segments_on_d3_cable_order_by_skip_count()
    if _1st_segments_on_d3_cables is not None and not _1st_segments_on_d3_cables.empty:
        for _d3_idx, _1st_segment in _1st_segments_on_d3_cables.iterrows():
            skip_count = int(_1st_segment[CABLE_SKIP_COUNT_FIELD_NAME])
            start_row_no = data_1st_row + skip_count
            fill_segment_and_next_segment_on_d3(ws_sro, start_row_no, start_col + 5, _1st_segment)
    # boxs_on_section = data_service_box.get_all_boxs_on_section_by_orders(section=section,
    #                                                                      sort_by=[BOX_IN_START_FIELD_NAME])
    # if boxs_on_section is not None and not boxs_on_section.empty:
    #     for _box_idx, _box in boxs_on_section.iterrows():
    #         fill_closure_port_on_section(ws_sro, data_1st_row, start_col + 5, _box)


def fill_segment_and_next_segment_on_d3(ws_sro, start_row_no, start_col, segment):
    if segment is None or segment.empty:
        return
    section = segment[CABLE_SECTION_FIELD_NAME]
    _type = segment[CABLE_TYPE_FIELD_NAME]
    # skip_count = int(segment[CABLE_SKIP_COUNT_FIELD_NAME])
    # port_start = int(segment[CABLE_PORT_START_FIELD_NAME])
    # port_end = int(segment[CABLE_PORT_END_FIELD_NAME])
    extremity_box = data_service_box.get_box_by_code(segment[CABLE_EXTREMITY_FIELD_NAME])
    in_start = extremity_box[BOX_IN_START_FIELD_NAME]
    in_end = extremity_box[BOX_IN_END_FIELD_NAME]

    fiber_to_fill_amt = in_end - in_start + 1
    row_no = start_row_no
    """填充d3线缆数据"""
    for i in range(0, fiber_to_fill_amt):
        row_no = start_row_no + i
        no = in_start + i
        code_cell = ws_sro.cell(row=row_no, column=start_col)
        code_cell.value = section

        type_cell = ws_sro.cell(row=row_no, column=start_col + 1)
        type_cell.value = _type

        no_cell = ws_sro.cell(row=row_no, column=start_col + 2)
        no_cell.value = no

        t = int((no - 1) / 12) + 1
        t_cell = ws_sro.cell(row=row_no, column=start_col + 3)
        t_cell.value = t
        t_cell.fill = PatternFill(fill_type="solid", start_color=get_fiber_color(t - 1))

        f = (no - 1) % 12 + 1
        f_cell = ws_sro.cell(row=row_no, column=start_col + 4)
        f_cell.value = f
        f_cell.fill = PatternFill(fill_type="solid", start_color=get_fiber_color(f - 1))

        f_cell = ws_sro.cell(row=row_no, column=start_col + 5)
        f_cell.value = f"{extremity_box[BOX_CODE_FIELD_NAME]}-{'%02d' % f}"
        
    next_segment = data_service_cable.get_next_segment(segment=segment)
    fill_segment_and_next_segment_on_d3(ws_sro, row_no + 1, start_col, next_segment)


def fill_d2_d3_fiber_data(ws_sro, data_1st_row, start_col):
    d2_d3_cables = data_service_cable.get_all_d2_d3_cables_order_by_skip_count()
    if d2_d3_cables is not None and not d2_d3_cables.empty:
        for _idx, cable in d2_d3_cables.iterrows():
            cable_level = cable[CABLE_LEVEL_FIELD_NAME]
            cable_start_col = start_col if cable_level == 2 else start_col + 10
            section = cable[CABLE_SECTION_FIELD_NAME]
            _type = cable[CABLE_TYPE_FIELD_NAME]
            skip_count = int(cable[CABLE_SKIP_COUNT_FIELD_NAME])
            port_start = int(cable[CABLE_PORT_START_FIELD_NAME])
            port_end = int(cable[CABLE_PORT_END_FIELD_NAME])
            for i in range(skip_count, skip_count + port_end - port_start + 1):
                code_cell = ws_sro.cell(row=data_1st_row + i, column=cable_start_col)
                code_cell.value = section

                type_cell = ws_sro.cell(row=data_1st_row + i, column=cable_start_col + 1)
                type_cell.value = _type

                on_cell = ws_sro.cell(row=data_1st_row + i, column=cable_start_col + 2)
                on_cell.value = i - skip_count + 1

                t = int((i - skip_count - 0.01) / 12) + 1
                t_cell = ws_sro.cell(row=data_1st_row + i, column=cable_start_col + 3)
                t_cell.value = t
                t_cell.fill = PatternFill(fill_type="solid", start_color=get_fiber_color(t))

                f = (i - skip_count - 1) % 12 + 1
                f_cell = ws_sro.cell(row=data_1st_row + i, column=cable_start_col + 4)
                f_cell.value = f
                f_cell.fill = PatternFill(fill_type="solid", start_color=get_fiber_color(f))

            boxs_on_section = data_service_box.get_all_boxs_on_section_by_orders(section=section,
                                                                                 sort_by=[BOX_IN_START_FIELD_NAME])
            if boxs_on_section is not None and not boxs_on_section.empty:
                for _box_idx, _box in boxs_on_section.iterrows():
                    fill_closure_port_on_section(ws_sro, data_1st_row, cable_start_col + 5, _box)

    return


def create_return_topo_cell(ws_sro):
    overview_cell = ws_sro.cell(row=1, column=1)
    overview_cell.value = "返回拓扑总览"
    overview_cell.hyperlink = f"#'拓扑总览'!A1"  # 直接链接到拓扑总览
    overview_cell.font = Font(color="0000FF", size=11, underline="single")
    overview_cell.alignment = TITLE_BASE_ALIGN


def draw_sro_table_header(ws_sro):
    # 2. 基础配置参数
    title_1st_row = 2  # 表头第一行（SRO、Splice State等所在行）
    start_col = 1  # 起始列（A列，对应数字1）
    font_name = "宋体"  # 全局字体

    for group_idx in range(1, 4):
        draw_title_group(ws_sro, font_name, group_idx, start_col, title_1st_row)
        start_col += COL_LOOP

    draw_rest_title(ws_sro, font_name, start_col, title_1st_row)

    # 调整行高（第2行和第3行，适配合并单元格内容）
    ws_sro.row_dimensions[1].height = 45
    ws_sro.row_dimensions[title_1st_row].height = 15.6  # 表头第一行（SRO等）
    ws_sro.row_dimensions[title_1st_row + 1].height = 15.6  # 表头第二行（SRO Port等）
    return title_1st_row + 2


def draw_rest_title(ws_sro, font_name, start_col, title_1st_row):
    cell_PBO = ws_sro.cell(row=title_1st_row, column=start_col)
    cell_PBO.value = "PBO"
    cell_PBO.font = Font(name=font_name, size=12, bold=True)
    cell_PBO.fill = PatternFill(fill_type="solid", start_color=TITLE_BOX_BG_COLOR)
    cell_PBO.border = CELL_BORDER
    cell_PBO.alignment = TITLE_BASE_ALIGN

    cell_PBO_PORT = ws_sro.cell(row=title_1st_row + 1, column=start_col)
    cell_PBO_PORT.value = "Port"
    cell_PBO_PORT.font = Font(name=font_name, size=11, bold=True)
    cell_PBO_PORT.fill = PatternFill(fill_type="solid", start_color=TITLE_BOX_BG_COLOR)
    cell_PBO_PORT.border = CELL_BORDER
    cell_PBO_PORT.alignment = TITLE_BASE_ALIGN

    cell_S = ws_sro.cell(row=title_1st_row, column=start_col + 1)
    cell_S.value = "S:"
    cell_S.font = Font(name=font_name, size=11, bold=True)
    cell_S.border = CELL_BORDER
    cell_S.alignment = TITLE_BASE_ALIGN

    cell_R = ws_sro.cell(row=title_1st_row + 1, column=start_col + 1)
    cell_R.value = "R:"
    cell_R.font = Font(name=font_name, size=11, bold=True)
    cell_R.border = CELL_BORDER
    cell_R.alignment = TITLE_BASE_ALIGN

    cell_Splice = ws_sro.cell(row=title_1st_row, column=start_col + 2)
    cell_Splice.value = "Splice"
    cell_Splice.font = Font(name=font_name, size=11)
    cell_Splice.border = CELL_BORDER
    cell_Splice.alignment = TITLE_BASE_ALIGN

    cell_Reserve = ws_sro.cell(row=title_1st_row + 1, column=start_col + 2)
    cell_Reserve.value = "Reserve"
    cell_Reserve.font = Font(name=font_name, size=11)
    cell_Reserve.border = CELL_BORDER
    cell_Reserve.alignment = TITLE_BASE_ALIGN


def draw_title_group(ws_sro, font_name, group_idx, start_col, title_1st_row):
    section_class = gen_section_class(group_idx)
    # 3. 第一组表头（A2:C2 + A3:C3）
    # ----------------------
    # 3.1 合并A2:C2（同一行，列：start_col ~ start_col+2）
    create_merged_cell(
        ws=ws_sro,
        start_row=title_1st_row,
        start_col=start_col,
        end_row=title_1st_row,
        end_col=start_col + 2,
        value="SRO",
        font=Font(name=font_name, size=12, bold=True, color="FFFFFF"),
        fill=PatternFill(fill_type="solid", start_color=TITLE_BOX_BG_COLOR),
        alignment=TITLE_BASE_ALIGN
    )
    # 设置A2单元格样式（合并后以左上角单元格为准）
    cell_a2 = ws_sro.cell(row=title_1st_row, column=start_col)
    cell_a2.value = "SRO" if group_idx == 1 else "Closure"
    cell_a2.font = Font(name=font_name, size=12, bold=True)  # 白色文字对比绿色背景
    cell_a2.fill = PatternFill(fill_type="solid", start_color=TITLE_BOX_BG_COLOR)
    cell_a2.border = CELL_BORDER
    cell_a2.alignment = TITLE_BASE_ALIGN
    # 3.2 第一组第二行（A3:C3）
    a3_c3_values = ["SRO Port", "ODF Code", "ODF Port"] if group_idx == 1 else ["Code", "Type", "NO."]  # 对应A3、B3、C3的值
    for col_offset in range(3):  # 循环处理A3（offset=0）、B3（offset=1）、C3（offset=2）
        cell = ws_sro.cell(
            row=title_1st_row + 1,
            column=start_col + col_offset
        )
        cell.value = a3_c3_values[col_offset]
        cell.font = Font(name=font_name, size=11, bold=True)
        cell.fill = PatternFill(fill_type="solid", start_color=TITLE_BOX_BG_COLOR)
        cell.border = CELL_BORDER
        cell.alignment = TITLE_BASE_ALIGN
    # ----------------------
    # 4. 第二组表头（D2:D3 + E2 + E3）- 含自动换行
    # ----------------------
    # 4.1 合并D2:D3（同一列，行：title_1st_row ~ title_1st_row+1）
    create_merged_cell(
        ws=ws_sro,
        start_row=title_1st_row,
        start_col=start_col + 3,
        end_row=title_1st_row + 1,
        end_col=start_col + 3,
        value="Splice State",
        font=Font(name=font_name, size=12, bold=True),
        fill=PatternFill(fill_type="solid", start_color=TITLE_SPLICE_BG_COLOR),
        alignment=TITLE_WRAP_ALIGN
    )
    # 设置D2单元格样式（合并后以左上角单元格为准，启用自动换行）
    cell_d2 = ws_sro.cell(row=title_1st_row, column=start_col + 3)
    cell_d2.value = "Splice State"  # 内容过长时会自动换行（也可手动加\n强制换行）
    cell_d2.font = Font(name=font_name, size=12, bold=True)
    cell_d2.fill = PatternFill(fill_type="solid", start_color=TITLE_SPLICE_BG_COLOR)
    cell_d2.border = CELL_BORDER
    cell_d2.alignment = TITLE_WRAP_ALIGN  # 关键：应用自动换行对齐
    # 4.2 处理E2和E3（空白单元格，仅保留样式）
    for row_offset in [0, 1]:  # 循环处理E2（row_offset=0）、E3（row_offset=1）
        cell = ws_sro.cell(
            row=title_1st_row + row_offset,
            column=start_col + 4
        )
        cell.fill = PatternFill(fill_type="solid", start_color=TITLE_BLANK_BG_COLOR)
        cell.border = CELL_BORDER
        cell.alignment = TITLE_BASE_ALIGN
    # ----------------------
    # 5. 第三组表头（F2:J2 + F3:J3）
    # ----------------------
    # 5.1 合并F2:J2（同一行，列：start_col+5 ~ start_col+9）
    create_merged_cell(
        ws=ws_sro,
        start_row=title_1st_row,
        start_col=start_col + 5,
        end_row=title_1st_row,
        end_col=start_col + 9,
        value="Distribution 1",
        font=Font(name=font_name, size=12, bold=True, color="FFFFFF"),
        fill=PatternFill(fill_type="solid", start_color=TITLE_SECTION_BG_COLOR),
        alignment=TITLE_BASE_ALIGN
    )
    # 设置F2单元格样式
    cell_f2 = ws_sro.cell(row=title_1st_row, column=start_col + 5)
    cell_f2.value = section_class
    cell_f2.font = Font(name=font_name, size=12, bold=True)  # 白色文字对比蓝色背景
    cell_f2.fill = PatternFill(fill_type="solid", start_color=TITLE_SECTION_BG_COLOR)
    cell_f2.border = CELL_BORDER
    cell_f2.alignment = TITLE_BASE_ALIGN
    # 5.2 第三组第二行（F3:J3）
    f3_j3_values = ["Code", "Type", "NO.", "T", "F"]  # 对应F3、G3、H3、I3、J3的值
    for col_offset in range(5):  # 循环处理5列
        cell = ws_sro.cell(
            row=title_1st_row + 1,
            column=start_col + 5 + col_offset
        )
        cell.value = f3_j3_values[col_offset]
        cell.font = Font(name=font_name, size=11, bold=True)
        cell.fill = PatternFill(fill_type="solid", start_color=TITLE_SECTION_BG_COLOR)
        cell.border = CELL_BORDER
        cell.alignment = TITLE_BASE_ALIGN
    # ----------------------
    # 6. 优化列宽和行高（避免内容溢出）
    # ----------------------
    # 调整列宽（A-J列，根据内容长度适配）
    col_widths = [20, 10, 10, 8, 4, 8, 8, 6, 6, 6]  # A到J列的宽度
    for col_idx in range(0, len(col_widths)):  # 1=A列，10=J列
        ws_sro.column_dimensions[excel_utils.num_to_col(col_idx + start_col)].width = col_widths[col_idx]


def gen_section_class(_class: int):
    return f"Distribution {_class}"


def create_merged_cell(ws, start_row, start_col, end_row, end_col, value, font, fill, alignment):
    """创建合并单元格并应用样式和边框"""
    # 合并单元格
    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=end_row, end_column=end_col)
    # 设置内容和样式（仅需设置左上角单元格）
    cell = ws.cell(row=start_row, column=start_col)
    cell.value = value
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    # 设置四周边框（关键：遍历所有单元格设置边缘）
    set_merged_border(ws, start_row, start_col, end_row, end_col)


def set_merged_border(ws, start_row, start_col, end_row, end_col):
    """为合并区域设置完整四周边框"""
    # 遍历合并区域的所有单元格
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            # 初始化边框
            border = Border()

            # 顶部边框：只给第一行设置上边框
            if row == start_row:
                border.top = BORDER_SIDE
            # 底部边框：只给最后一行设置下边框
            if row == end_row:
                border.bottom = BORDER_SIDE
            # 左侧边框：只给第一列设置左边框
            if col == start_col:
                border.left = BORDER_SIDE
            # 右侧边框：只给最后一列设置右边框
            if col == end_col:
                border.right = BORDER_SIDE

            cell.border = border


COL_LOOP = 10
ODF_MAX_PORT_NO = 144
TARGET_ROW = 480

# 对齐方式配置（水平居中+垂直居中，第二组额外加自动换行）
TITLE_BASE_ALIGN = Alignment(horizontal="center", vertical="center")  # 基础对齐（无自动换行）
TITLE_WRAP_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 带自动换行的对齐

# 边框样式（统一黑色细边框）
BORDER_SIDE = Side(style="thin", color="000000")
CELL_BORDER = Border(
    left=BORDER_SIDE,
    right=BORDER_SIDE,
    top=BORDER_SIDE,
    bottom=BORDER_SIDE
)
SHEET_ZOOM_SCALE = 70
# 颜色定义（RGB格式）
TITLE_BOX_BG_COLOR = "00B050"  # 第一组绿色背景
TITLE_SPLICE_BG_COLOR = "FFFF00"  # 第二组黄色背景
TITLE_BLANK_BG_COLOR = "FFFFFF"  # 第三组蓝色背景
TITLE_SECTION_BG_COLOR = "00B0F0"  # 第三组蓝色背景

FIBER_COLOR = {
    0: "FF0000",  # idx=1 红
    1: "0066FF",  # idx=2 蓝
    2: "00B050",  # idx=3 绿
    3: "FFFF00",  # idx=4 黄
    4: "7030A0",  # idx=5 紫
    5: "FFFFFF",  # idx=6 白
    6: "FFCC00",  # idx=7 金
    7: "D9D9D9",  # idx=8 灰
    8: "963634",  # idx=9 赭
    9: "FDE9D9",  # idx=10 米
    10: "66FFFF",  # idx=11 青
    11: "FFCCFF"  # idx=12 粉
}


def get_fiber_color(no: int):
    return FIBER_COLOR[no % 12]
