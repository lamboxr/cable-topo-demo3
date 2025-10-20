import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment

import init_data
from data_service import data_service_box, data_service_cable
from utils import excel_utils

# 全局配置
COLUMN_WIDTHS = {
    'A': 12,
    'B': 48,
    'C': 32,
    'D': 48,
    'E': 32,
    'F': 48,
    'G': 32
}
LEVEL_TO_COLUMN = {
    1: 'B',
    2: 'D',
    3: 'F'
}
GROUP_ROWS = 7  # 每组占用8行


def init_workbook():
    """初始化Excel工作簿，设置列宽和样式基础"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "逻辑拓扑图"

    # 设置列宽
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    return wb, ws


def generate_topo_excel(output_path):
    """生成拓扑Excel的主函数"""
    wb, ws = init_workbook()
    current_row = 2  # 起始行

    # 1. 获取所有SRO节点（根节点），按code升序
    sro_points = data_service_nap.get_all_sro_points_by_order_code_asc()
    if sro_points is None or sro_points.empty:
        print("无SRO节点数据，无法生成拓扑图")
        return

    # 2. 遍历SRO节点，递归绘制拓扑
    for _, sro in sro_points.iterrows():
        point_branches_start_row = current_row
        point_branches_end_row = point_branches_start_row
        # 绘制SRO节点（A列）
        current_row = draw_sro_node(ws, current_row, sro)
        # 查询SRO节点关联的线缆（Level1），按code升序
        d1_cables = data_service_cable.get_all_cables_start_with_one_point_order_by_code_asc(sro['code'])
        if d1_cables is not None and len(d1_cables) > 1:
            point_branches_start_row = current_row + 4
        if d1_cables is None or d1_cables.empty:
            current_row += GROUP_ROWS
            continue
        if d1_cables is not None and not d1_cables.empty:
            d1_cable_idx = 0
            d1_cables_size = len(d1_cables)
            for _, cable in d1_cables.iterrows():
                if d1_cable_idx == d1_cables_size - 1 and d1_cables_size > 1:
                    point_branches_end_row = current_row + 1
                # 绘制线缆（B列，Level1）并递归处理下一级节点
                current_row = draw_cable_and_recurse(ws=ws, start_row=current_row, cable_data=cable,
                                                     current_level=cable['level'],
                                                     upper_level=cable['level'] - 1)
                d1_cable_idx += 1
            if len(d1_cables) > 1:
                for row in range(point_branches_start_row, point_branches_end_row):
                    set_cell(
                        ws,
                        row=row,
                        col="A",
                        value="",
                        border=CABLE_ROUTE_BORDER
                    )

    # 保存文件
    wb.save(output_path)
    print(f"拓扑图已生成：{output_path}")


def draw_sro_node(ws, start_row, sro_data):
    """绘制A列的SRO节点（8行一组，不合并单元格）"""
    # 第1-4行：粗外侧线框（模拟盒子）
    for row in range(start_row, start_row + 4):  # 行索引：start_row到start_row+3
        # 第1行：nap.class（加粗居中）
        if row == start_row:
            set_cell(
                ws,
                row=row,
                col='A',
                value=sro_data['class'],
                border=BOX_1ST_ROW_BORDER,
                font=BOLD_FONT,
                align=CENTER_ALIGN
            )
        # 第2行：nap.code（居中）
        elif row == start_row + 1:
            set_cell(
                ws,
                row=row,
                col='A',
                value=sro_data['code'],
                border=BOX_MIDDLE_ROW_BORDER,
                align=CENTER_ALIGN
            )
        elif row == start_row + 2:
            set_cell(
                ws,
                row=row,
                col='A',
                value=None,
                border=BOX_MIDDLE_ROW_BORDER
            )
        # 第3-4行：留空（保持边框）
        else:
            set_cell(
                ws,
                row=row,
                col='A',
                value=None,
                border=BOX_LAST_ROW_BORDER
            )
    # 第5-8行：留空（无边框）
    for row in range(start_row + 4, start_row + GROUP_ROWS):
        set_cell(ws, row=row, col='A', value=None)
    return start_row  # 移动到下一组


def draw_closure_pbo_node(ws, start_row, nap_data, col):
    """绘制C/E/G列的Closure/PBO节点（8行一组，不合并单元格）"""
    # 第1-4行：粗外侧线框（模拟盒子）
    for row in range(start_row, start_row + 4):
        # 第1行：nap.class（加粗居中）
        if row == start_row:
            set_cell(
                ws,
                row=row,
                col=col,
                value=nap_data['class'],
                border=BOX_1ST_ROW_BORDER,
                font=BOLD_FONT,
                align=CENTER_ALIGN
            )
        # 第2行：nap.code + "    " + nap.type（居中）
        elif row == start_row + 1:
            set_cell(
                ws,
                row=row,
                col=col,
                value=f"{nap_data['code']}    {nap_data['type']}",
                border=BOX_MIDDLE_ROW_BORDER,
                align=CENTER_ALIGN
            )
        # 第3行：留空
        elif row == start_row + 2:
            set_cell(
                ws,
                row=row,
                col=col,
                value=f"On:{nap_data['cable_in']}",
                border=BOX_MIDDLE_ROW_BORDER,
                align=CENTER_ALIGN
            )
        # 第4行：nap.in_start + "-" + nap.in_end（居中）
        elif row == start_row + 3:
            set_cell(
                ws,
                row=row,
                col=col,
                value=f"InRange:{int(nap_data['in_start'])}-{int(nap_data['in_end'])}",
                border=BOX_LAST_ROW_BORDER,
                align=CENTER_ALIGN
            )
    # 第5-8行：留空（无边框）
    for row in range(start_row + 4, start_row + GROUP_ROWS):
        set_cell(ws, row=row, col=col, value=None)
    return start_row  # 移动到下一组


def draw_cable(ws, start_row, cable_data, current_level, upper_level, draw_cable_right_border):
    """绘制B/D/F列的线缆（8行一组，不合并单元格）"""
    # 根据level确定列（B=1, D=2, F=3）
    col = LEVEL_TO_COLUMN[current_level]
    # level显示文本映射
    level_text = {1: "Distribution 01", 2: "Distribution 02", 3: "Distribution 03"}[current_level]

    if current_level - upper_level > 1:
        set_cell(
            ws,
            row=start_row,
            col=excel_utils.get_left_col_letter(col),
            value="",
            border=CABLE_FIRST_ROW_BORDER
        )
        set_cell(
            ws,
            row=start_row,
            col=excel_utils.get_left_col_letter(col, 2),
            value="",
            border=CABLE_FIRST_ROW_BORDER
        )
    route_border = Border()
    if draw_cable_right_border:
        route_border = CABLE_ROUTE_BORDER

    # 第1行：level文本（下边框加粗）
    set_cell(
        ws,
        row=start_row,
        col=col,
        value=level_text,
        border=CABLE_FIRST_ROW_BORDER,
        font=BOLD_FONT,
        align=CENTER_ALIGN
    )
    # 第2行：cable.code + "    " + cable.type
    set_cell(
        ws,
        row=start_row + 1,
        col=col,
        value=f"{cable_data['code']}    {cable_data['type']}",
        align=CENTER_ALIGN,
        border=route_border
    )
    # 第3行：cable.r_nodes
    set_cell(
        ws,
        row=start_row + 2,
        col=col,
        value=f"From: {cable_data['origin_box']}    RNodes:{cable_data['r_nodes']}",
        align=CENTER_ALIGN,
        border=route_border
    )
    # 第4行：cable.port_start + "-" + cable.port_end
    set_cell(
        ws,
        row=start_row + 3,
        col=col,
        value=f"PortRange:{int(cable_data['port_start'])}-{int(cable_data['port_end'])}",
        align=CENTER_ALIGN,
        border=route_border
    )
    # 第5-8行：留空
    for row in range(start_row + 4, start_row + GROUP_ROWS):
        set_cell(ws, row=row, col=col, value=None, border=route_border)
    return start_row  # 移动到下一组


def draw_cable_and_recurse(ws, start_row, cable_data, current_level, upper_level):
    """绘制线缆并递归处理下一级节点"""
    draw_cable_right_border = False

    # 1. 查询线缆上的点（按pass_seq顺序：中间点1-99 → 终点100）
    points_on_cable = data_service_nap.get_all_points_on_cable_by_orders(
        cable_code=cable_data['code'],
        sort_by=['pass_seq'],
        ascending=True
    )

    # 2. 判断线缆下的点大于1,线缆就要描绘2-7的右边框
    if len(points_on_cable) > 1:
        draw_cable_right_border = True
    # 2. 绘制当前线缆（占用8行）
    current_row = draw_cable(ws, start_row, cable_data, current_level, upper_level, draw_cable_right_border)

    route_border_start_row = current_row
    route_border_end_row = route_border_start_row

    if points_on_cable is None or points_on_cable.empty:
        return current_row + GROUP_ROWS
    elif len(points_on_cable) > 1:
        route_border_start_row = current_row + GROUP_ROWS
        current_row += GROUP_ROWS

    # 3. 确定下一级节点的列（线缆列的右侧列：B→C，D→E，F→G）
    next_col_map = {'B': 'C', 'D': 'E', 'F': 'G'}
    next_col = next_col_map[LEVEL_TO_COLUMN[current_level]]

    # 4. 遍历点并绘制（每个点占用8行）
    point_idx = 0
    points_size = len(points_on_cable)
    for _, point in points_on_cable.iterrows():
        point_branches_start_row = current_row
        point_branches_end_row = point_branches_start_row
        if point_idx == points_size - 1 and points_size > 1:
            route_border_end_row = current_row
        # 根据点类型绘制节点（Closure/PBO）
        if point['class'] in ['Closure', 'PBO']:
            current_row = draw_closure_pbo_node(ws, current_row, point, next_col)

            # 5. 递归查询该点作为起点的下一级线缆（level+1）
            # next_level = level + 1
            # if next_level <= 3:  # 最多到level 3
            sub_cables = data_service_cable.get_all_cables_start_with_one_point_order_by_code_asc(
                nap_code=point['code']
            )
            if sub_cables is not None and len(sub_cables) > 1:
                point_branches_start_row = current_row + 4
            if sub_cables is None or sub_cables.empty:
                current_row += GROUP_ROWS
            elif sub_cables is not None and not sub_cables.empty:
                sub_cable_idx = 0
                sub_cables_size = len(sub_cables)
                for __, next_cable in sub_cables.iterrows():
                    if sub_cable_idx == sub_cables_size - 1 and sub_cables_size > 1:
                        point_branches_end_row = current_row + 1
                    current_row = draw_cable_and_recurse(ws, current_row, next_cable, next_cable['level'],
                                                         current_level)
                    sub_cable_idx += 1
                if len(sub_cables) > 1:
                    for row in range(point_branches_start_row, point_branches_end_row):
                        set_cell(
                            ws,
                            row=row,
                            col=next_col,
                            value="",
                            border=CABLE_ROUTE_BORDER
                        )
            point_idx += 1

    if len(points_on_cable) > 1:
        for row in range(route_border_start_row, route_border_end_row):
            set_cell(
                ws,
                row=row,
                col=excel_utils.get_left_col_letter(next_col),
                value="",
                border=CABLE_ROUTE_BORDER
            )

    return current_row

THIN_WIDTH = 'thin'
THICKER_WIDTH = 'medium'

# 样式定义调整：为独立单元格设置完整边框
BOX_1ST_ROW_BORDER = Border(
    left=Side(style=THICKER_WIDTH),
    right=Side(style=THICKER_WIDTH),
    top=Side(style=THICKER_WIDTH),
    bottom=Side(style=THIN_WIDTH)
)

BOX_MIDDLE_ROW_BORDER = Border(
    left=Side(style=THICKER_WIDTH),
    right=Side(style=THICKER_WIDTH),
)

BOX_LAST_ROW_BORDER = Border(
    left=Side(style=THICKER_WIDTH),
    right=Side(style=THICKER_WIDTH),
    bottom=Side(style=THICKER_WIDTH)
)

# 线缆第一行下边框加粗（其他边框默认）
CABLE_FIRST_ROW_BORDER = Border(
    bottom=Side(style=THICKER_WIDTH)
)

CABLE_ROUTE_BORDER = Border(
    right=Side(style=THICKER_WIDTH)
)

POINT_BRANCHES_BORDER = Border(
    right=Side(style=THICKER_WIDTH)
)

BOLD_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)


def merge_cells(ws, start_row, end_row, col, value, border=None, font=None, align=None):
    """
    合并单元格并设置内容和样式
    :param ws: 工作表
    :param start_row: 起始行
    :param end_row: 结束行
    :param col: 列标识（如'A'）
    :param value: 单元格内容
    :param border: 边框样式
    :param font: 字体样式
    :param align: 对齐方式
    """
    cell_range = f"{col}{start_row}:{col}{end_row}"
    # ws.merge_cells(cell_range)
    cell = ws[f"{col}{start_row}"]
    cell.value = value
    cell.border = border or Border()
    cell.font = font or Font()
    cell.alignment = align or Alignment()


def set_cell(ws, row, col, value, border=None, font=None, align=None):
    """设置单个单元格内容和样式"""
    cell = ws[f"{col}{row}"]
    cell.value = value
    cell.border = border or Border()
    cell.font = font or Font()
    cell.alignment = align or Alignment()


if __name__ == '__main__':
    init_data.main()
    generate_topo_excel("拓扑图输出.xlsx")