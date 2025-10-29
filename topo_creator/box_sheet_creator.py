from dataclasses import fields

from openpyxl.worksheet.hyperlink import Hyperlink

from constraints.field_name_mapper import BOX_CODE_FIELD_NAME, CABLE_SKIP_COUNT_FIELD_NAME, CABLE_PORT_START_FIELD_NAME, \
    CABLE_PORT_END_FIELD_NAME, CABLE_CODE_FIELD_NAME, CABLE_TYPE_FIELD_NAME, CABLE_SECTION_FIELD_NAME, \
    BOX_IN_START_FIELD_NAME, BOX_IN_END_FIELD_NAME, BOX_TYPE_FIELD_NAME, CABLE_LEVEL_FIELD_NAME, \
    CABLE_EXTREMITY_FIELD_NAME
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
# from data_service.service_manager import global_services  # 导入全局服务
# from data_service import data_service_cable, data_service_box
from utils import excel_utils

FONT_NAME = '宋体'

class BoxSheetCreator:
    def __init__(self, data_services) -> None:
        # 仅初始化一次（使用类级别标记判断）
        # 初始化三个服务（通过third_party_params获取路径和图层名）
        self.sro_service = data_services.sro_service
        self.box_service = data_services.box_service
        self.cable_service = data_services.cable_service


    def create_box_sheet(self, ws, topo_box_col, box_data, sro_sheet_name):
        if topo_box_col == 'A':
            return self._create_sro_sheet(ws, box_data)
        elif topo_box_col == 'G':
            return self._create_pbo_sheet(ws, box_data, sro_sheet_name)
        return None


    def _create_pbo_sheet(self, ws, box_data, sro_sheet_name):
        pbo_code = box_data[BOX_CODE_FIELD_NAME]
        pbo_sheet_name = pbo_code
        wb = ws.parent
        if pbo_sheet_name not in wb.sheetnames:
            ws_box = wb.create_sheet(title=pbo_sheet_name)
            ws_box.sheet_view.zoomScale = SHEET_ZOOM_SCALE
            ws_box.append([])
            # 定义默认字体（例如：微软雅黑，11号，非粗体）
            default_font = Font(name=FONT_NAME, size=11)

            # 创建默认样式并关联字体
            default_style = NamedStyle(name="default_style", font=default_font)

            # 将默认样式应用到整个工作表（A1到XFD1048576，即所有单元格）
            ws_box.style = default_style
            start_col = BOX_SHEET_TABLE_START_COL
            # 绘制返回按钮
            self.create_return_topo_cell(ws_box, ws.title)
            # 绘制表头
            self.draw_box_sheet_title(ws_box, box_data[BOX_CODE_FIELD_NAME], start_col)
            data_1st_row = self.draw_box_sheet_header(ws_box, start_col)
            if sro_sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sro_sheet_name}' 不存在于workbook对象中")
            ws_sro = wb[sro_sheet_name]
            # 在AE列中查找以"PBO_NAME"开头的连续单元格区间
            row_begin = None
            row_end = None
            ae_col = start_col + excel_utils.col_to_num('AD')  # AE列对应的数字（A=1, B=2, ..., AE=31）
            # current_row = start_row_no
            row_count = ws_sro.max_row

            # 遍历AE列，找到起始行
            for current_row in range(data_1st_row, data_1st_row + row_count):
                cell_value = ws_sro.cell(row=current_row, column=ae_col).value
                if cell_value and str(cell_value).startswith(pbo_code):
                    row_begin = current_row
                    break

            # 找到连续区间的结束行
            for current_row in range(row_begin, row_begin + row_count):
                cell_value = ws_sro.cell(row=current_row, column=ae_col).value
                if not (cell_value and str(cell_value).startswith(pbo_code)):
                    row_end = current_row - 1
                    break

            # 复制A列到AE列（1到31列）的指定区间到目标工作簿
            self.copy_cells_with_style(
                source_ws=ws_sro,
                target_ws=ws_box,
                source_start_row=row_begin,
                source_end_row=row_end,
                source_start_col=1,  # A列
                source_end_col=ae_col,  # AE列
                target_start_row=data_1st_row,  # 目标开始行
                target_start_col=1  # 目标开始列（A列）
            )

            ws_box.freeze_panes = f"{excel_utils.num_to_col(BOX_SHEET_TABLE_START_COL)}4"
            return pbo_sheet_name

        return None


    def copy_cells_with_style(self, source_ws, target_ws, source_start_row, source_end_row,
                              source_start_col, source_end_col, target_start_row, target_start_col):
        """
        复制源工作表指定区域到目标工作表，包含值和所有样式
        """
        for row in range(source_start_row, source_end_row + 1):
            for col in range(source_start_col, source_end_col + 1):
                # 源单元格
                source_cell = source_ws.cell(row=row, column=col)

                # 计算目标单元格位置
                target_row = target_start_row + (row - source_start_row)
                target_col = target_start_col + (col - source_start_col)
                target_cell = target_ws.cell(row=target_row, column=target_col)

                # 复制值
                target_cell.value = source_cell.value

                # 复制字体样式
                if source_cell.font:
                    target_cell.font = Font(
                        name=FONT_NAME,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic,
                        color=None if target_row == 4 and target_col ==31 else source_cell.font.color,
                        underline=None if target_row == 4 and target_col ==31 else source_cell.font.underline
                    )

                # 复制填充样式
                if source_cell.fill:
                    target_cell.fill = PatternFill(
                        fill_type=source_cell.fill.fill_type,
                        start_color=source_cell.fill.start_color,
                        end_color=source_cell.fill.end_color
                    )

                # 复制边框样式
                if source_cell.border:
                    target_cell.border = Border(
                        left=Side(
                            border_style=source_cell.border.left.border_style,
                            color=source_cell.border.left.color
                        ),
                        right=Side(
                            border_style=source_cell.border.right.border_style,
                            color=source_cell.border.right.color
                        ),
                        top=Side(
                            border_style=source_cell.border.top.border_style,
                            color=source_cell.border.top.color
                        ),
                        bottom=Side(
                            border_style=source_cell.border.bottom.border_style,
                            color=source_cell.border.bottom.color
                        )
                    )

                # 复制对齐方式
                if source_cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical,
                        text_rotation=source_cell.alignment.text_rotation,
                        wrap_text=source_cell.alignment.wrap_text,
                        shrink_to_fit=source_cell.alignment.shrink_to_fit,
                        indent=source_cell.alignment.indent
                    )


    def _create_sro_sheet(self, ws, sro_data):
        wb = ws.parent
        sheet_name = sro_data[BOX_CODE_FIELD_NAME]
        # 获取工作簿对象（从当前工作表反向获取）

        # 检查子sheet是否已存在，不存在则创建
        if sheet_name not in wb.sheetnames:
            ws_sro = wb.create_sheet(title=sheet_name)
            ws_sro.sheet_view.zoomScale = SHEET_ZOOM_SCALE
            ws_sro.append([])
            # 定义默认字体（例如：微软雅黑，11号，非粗体）
            default_font = Font(name=FONT_NAME, size=11)

            # 创建默认样式并关联字体
            default_style = NamedStyle(name="default_style", font=default_font)

            # 将默认样式应用到整个工作表（A1到XFD1048576，即所有单元格）
            ws_sro.style = default_style
            start_col = BOX_SHEET_TABLE_START_COL
            # 绘制返回按钮
            self.create_return_topo_cell(ws_sro, ws.title)
            # 绘制表头
            self.draw_box_sheet_title(ws_sro, sro_data[BOX_CODE_FIELD_NAME], start_col)
            data_1st_row = self.draw_box_sheet_header(ws_sro, start_col)
            self.fill_d1_data(ws_sro, data_1st_row, start_col)
            self.fill_d2_data(ws_sro, data_1st_row, start_col + COL_LOOP + 5)
            self.fill_d3_data(ws_sro, data_1st_row, start_col + COL_LOOP * 2)
            self.change_splice_state(ws_sro,data_1st_row, start_col)
            ws_sro.freeze_panes = f"{excel_utils.num_to_col(start_col)}{data_1st_row}"
            return sheet_name
        return None

    def _change_splice_state(self, ws_sro, data_1st_row, start_col):
        _2nd_splice_col = start_col + COL_LOOP +3
        _3rd_splice_col = _2nd_splice_col + COL_LOOP
        d2_start_row = _2nd_splice_col + 2
        d3_start_row = _3rd_splice_col +2
        current_row = data_1st_row
        while True:
            _2nd_splice_cell = ws_sro.cell(row=current_row, column=_2nd_splice_col)
            if not _2nd_splice_cell.value:
                break
            ws_sro[f"{excel_utils.num_to_col(_2nd_splice_col)}{current_row}"] = f'=IF(AND(ISBLANK({excel_utils.num_to_col(d2_start_row)}{current_row}),ISBLANK({excel_utils.num_to_col(d3_start_row)}{current_row})),"R","S")'

            _3rd_splice_cell = ws_sro.cell(row=current_row, column=_3rd_splice_col)
            if _3rd_splice_cell.value:
                ws_sro[f"{excel_utils.num_to_col(_3rd_splice_col)}{current_row}"] = f'=IF(ISBLANK({excel_utils.num_to_col(d3_start_row)}{current_row}),"R","S")'
            current_row += 1

    def change_splice_state(self, ws_sro, data_1st_row, start_col):
        _2nd_splice_col = start_col + 3 + COL_LOOP
        _3rd_splice_col = _2nd_splice_col+ COL_LOOP
        current_row = data_1st_row - 1
        while True:
            current_row += 1
            _2nd_splice_cell = ws_sro.cell(row=current_row, column=_2nd_splice_col)
            d2_empty = True
            d3_empty = True
            if not _2nd_splice_cell.value:
                break
            for i in range(_2nd_splice_col+1, _2nd_splice_col+6):
                if ws_sro.cell(row=current_row, column=start_col+i).value:
                    d2_empty = False
                    break


            for i in range(_3rd_splice_col + 1, _3rd_splice_col + 6):
                if ws_sro.cell(row=current_row, column=i).value:
                    d3_empty = False
                    break
            if d2_empty and d3_empty:
                ws_sro[f"{excel_utils.num_to_col(_2nd_splice_col)}{current_row}"] = "R"

            _3rd_splice_cell = ws_sro.cell(row=current_row, column=_3rd_splice_col)
            if _3rd_splice_cell.value and d3_empty:
                ws_sro[f"{excel_utils.num_to_col(_3rd_splice_col)}{current_row}"] = "R"


    def fill_d1_data(self, ws_sro, data_1st_row, start_col):
        _1st_segments = self.cable_service.get_all_1st_segments_on_d1_section_order_by_skip_count_asc()
        # 初始化数据
        if _1st_segments is not None and not _1st_segments.empty:
            sro_port_idx = 1
            odf_code_idx = 1
            odf_port_idx = 1
            for _idx, _1st_segment in _1st_segments.iterrows():
                skip_count = int(_1st_segment[CABLE_SKIP_COUNT_FIELD_NAME])
                section = _1st_segment[CABLE_SECTION_FIELD_NAME]
                port_start = int(_1st_segment[CABLE_PORT_START_FIELD_NAME])
                port_end = int(_1st_segment[CABLE_PORT_END_FIELD_NAME])

                fiber_to_fill_amt = port_end - port_start + 1
                for i in range(0, fiber_to_fill_amt):
                    row_no = data_1st_row + skip_count + i
                    #sro_port
                    cell = ws_sro.cell(row=row_no, column=start_col)
                    cell.value = sro_port_idx
                    cell.alignment = CENTER_ALIGN

                    # odf_code_
                    cell = ws_sro.cell(row=row_no, column=start_col + 1)
                    cell.value = f"ODF{'%02d' % odf_code_idx}"
                    cell.alignment = CENTER_ALIGN

                    # odf_port_
                    cell = ws_sro.cell(row=row_no, column=start_col + 2)
                    cell.value = odf_port_idx
                    cell.alignment = CENTER_ALIGN

                    # splice
                    cell = ws_sro.cell(row=row_no, column=start_col + 3)
                    cell.value = 'S'
                    cell.alignment = CENTER_ALIGN

                    # cable_code_
                    cell = ws_sro.cell(row=row_no, column=start_col + 5)
                    cell.value = section
                    cell.alignment = CENTER_ALIGN

                    # cable_type_
                    cell = ws_sro.cell(row=row_no, column=start_col + 6)
                    cell.value = _1st_segment[CABLE_TYPE_FIELD_NAME]
                    cell.alignment = CENTER_ALIGN

                    # cable_no_
                    cell = ws_sro.cell(row=row_no, column=start_col + 7)
                    cell.value = i + 1
                    cell.alignment = CENTER_ALIGN

                    # cable_t_
                    cell = ws_sro.cell(row=row_no, column=start_col + 8)
                    t = int(i / 12) + 1
                    cell.value = t
                    cell.font = Font(color=get_fiber_font_color(t - 1), bold=True)
                    bg_color = get_fiber_bg_color(t - 1)
                    if bg_color:
                       cell.fill = PatternFill(fill_type="solid", start_color=bg_color)
                    cell.alignment = CENTER_ALIGN

                    # cable_f_
                    cell = ws_sro.cell(row=row_no, column=start_col + 9)
                    f = i % 12 + 1
                    cell.value = f
                    cell.font = Font(color=get_fiber_font_color(f - 1))
                    bg_color = get_fiber_bg_color(f - 1)
                    if bg_color:
                        cell.fill = PatternFill(fill_type="solid", start_color=bg_color)
                    cell.alignment = CENTER_ALIGN

                    sro_port_idx += 1
                    odf_code_idx = odf_code_idx if odf_port_idx < ODF_MAX_PORT_NO else odf_code_idx + 1
                    odf_port_idx = odf_port_idx + 1 if odf_port_idx < ODF_MAX_PORT_NO else 1

                boxs_on_section = self.box_service.get_all_boxs_on_section_by_orders(section=section,
                                                                                     sort_by=[BOX_IN_START_FIELD_NAME])
                if boxs_on_section is not None and not boxs_on_section.empty:
                    for _box_idx, _box in boxs_on_section.iterrows():
                        self.fill_closure_port_on_section(ws_sro, data_1st_row, start_col + COL_LOOP, _box)


    def fill_d2_data(self, ws_sro, data_1st_row, start_col):
        _1st_segments_on_d2_section = self.cable_service.get_all_1st_segments_on_d2_section_order_by_skip_count_asc()
        if _1st_segments_on_d2_section is not None and not _1st_segments_on_d2_section.empty:
            for _idx, _1st_segment in _1st_segments_on_d2_section.iterrows():
                skip_count = int(_1st_segment[CABLE_SKIP_COUNT_FIELD_NAME])
                section = _1st_segment[CABLE_SECTION_FIELD_NAME]
                port_start = int(_1st_segment[CABLE_PORT_START_FIELD_NAME])
                port_end = int(_1st_segment[CABLE_PORT_END_FIELD_NAME])
                fiber_to_fill_amt = port_end - port_start + 1
                for i in range(0, fiber_to_fill_amt):
                    no = i + 1
                    row_no = data_1st_row + skip_count + i
                    cell = ws_sro.cell(row=row_no, column=start_col)
                    cell.value = section
                    cell.alignment = CENTER_ALIGN

                    cell = ws_sro.cell(row=row_no, column=start_col + 1)
                    cell.value = _1st_segment[CABLE_TYPE_FIELD_NAME]
                    cell.alignment = CENTER_ALIGN

                    cell = ws_sro.cell(row=row_no, column=start_col + 2)
                    cell.value = no
                    cell.alignment = CENTER_ALIGN

                    t = int(i / 12) + 1
                    cell = ws_sro.cell(row=row_no, column=start_col + 3)
                    cell.value = t
                    cell.font = Font(color=get_fiber_font_color(t - 1),bold=True)
                    bg_color = get_fiber_bg_color(t - 1)
                    if bg_color:
                        cell.fill = PatternFill(fill_type="solid", start_color=bg_color)
                    cell.alignment = CENTER_ALIGN

                    f = i % 12 + 1
                    cell = ws_sro.cell(row=row_no, column=start_col + 4)
                    cell.value = f
                    cell.font = Font(color=get_fiber_font_color(f - 1))
                    bg_color = get_fiber_bg_color(f - 1)
                    if bg_color:
                        cell.fill = PatternFill(fill_type="solid", start_color=bg_color)
                    cell.alignment = CENTER_ALIGN

                boxs_on_section = self.box_service.get_all_boxs_on_section_by_orders(section=section,
                                                                                     sort_by=[BOX_IN_START_FIELD_NAME])
                if boxs_on_section is not None and not boxs_on_section.empty:
                    for _box_idx, _box in boxs_on_section.iterrows():
                        self.fill_closure_port_on_section(ws_sro, data_1st_row, start_col + 5, _box)


    def fill_d3_data(self, ws_sro, data_1st_row, start_col):
        _1st_segments_on_d3_cables = self.cable_service.get_all_1st_segments_on_d3_cable_order_by_skip_count()
        if _1st_segments_on_d3_cables is not None and not _1st_segments_on_d3_cables.empty:
            for _d3_idx, _1st_segment in _1st_segments_on_d3_cables.iterrows():
                skip_count = int(_1st_segment[CABLE_SKIP_COUNT_FIELD_NAME])
                start_row_no = data_1st_row + skip_count
                self.fill_segment_and_next_segment_on_d3(ws_sro, start_row_no, start_col + 5, _1st_segment)
        # boxs_on_section = data_service_box.get_all_boxs_on_section_by_orders(section=section,
        #                                                                      sort_by=[BOX_IN_START_FIELD_NAME])
        # if boxs_on_section is not None and not boxs_on_section.empty:
        #     for _box_idx, _box in boxs_on_section.iterrows():
        #         fill_closure_port_on_section(ws_sro, data_1st_row, start_col + 5, _box)


    def fill_closure_port_on_section(self, ws_sro, data_1st_row, start_col, box_data):
        in_start = box_data[BOX_IN_START_FIELD_NAME]
        in_end = box_data[BOX_IN_END_FIELD_NAME]
        skip_count = int(box_data[CABLE_SKIP_COUNT_FIELD_NAME])
        port_to_fill_amt = in_end - in_start + 1
        for i in range(0, port_to_fill_amt):
            no = i + 1
            row_no = data_1st_row + skip_count + i
            cell = ws_sro.cell(row=row_no, column=start_col)
            cell.value = box_data[BOX_CODE_FIELD_NAME]
            cell.alignment = CENTER_ALIGN

            cell = ws_sro.cell(row=row_no, column=start_col + 1)
            cell.value = box_data[BOX_TYPE_FIELD_NAME]
            cell.alignment = CENTER_ALIGN

            cell = ws_sro.cell(row=row_no, column=start_col + 2)
            cell.value = no
            cell.alignment = CENTER_ALIGN

            cell = ws_sro.cell(row=row_no, column=start_col + 3)
            cell.value = 'S'
            cell.alignment = CENTER_ALIGN


    def fill_segment_and_next_segment_on_d3(self, ws_sro, start_row_no, start_col, segment):
        if segment is None or segment.empty:
            return
        section = segment[CABLE_SECTION_FIELD_NAME]
        _type = segment[CABLE_TYPE_FIELD_NAME]
        # skip_count = int(segment[CABLE_SKIP_COUNT_FIELD_NAME])
        # port_start = int(segment[CABLE_PORT_START_FIELD_NAME])
        # port_end = int(segment[CABLE_PORT_END_FIELD_NAME])
        extremity_box = self.box_service.get_box_by_code(segment[CABLE_EXTREMITY_FIELD_NAME])
        in_start = extremity_box[BOX_IN_START_FIELD_NAME]
        in_end = extremity_box[BOX_IN_END_FIELD_NAME]

        fiber_to_fill_amt = in_end - in_start + 1
        row_no = start_row_no
        """填充d3线缆数据"""
        for i in range(0, fiber_to_fill_amt):
            row_no = start_row_no + i
            no = in_start + i
            cell = ws_sro.cell(row=row_no, column=start_col)
            cell.value = section
            cell.alignment = CENTER_ALIGN

            cell = ws_sro.cell(row=row_no, column=start_col + 1)
            cell.value = _type
            cell.alignment = CENTER_ALIGN

            cell = ws_sro.cell(row=row_no, column=start_col + 2)
            cell.value = no
            cell.alignment = CENTER_ALIGN

            t = int((no - 1) / 12) + 1
            cell = ws_sro.cell(row=row_no, column=start_col + 3)
            cell.value = t
            cell.alignment = CENTER_ALIGN
            cell.font = Font(color=get_fiber_font_color(t - 1), bold=True)
            bg_color = get_fiber_bg_color(t - 1)
            if bg_color:
                cell.fill = PatternFill(fill_type="solid", start_color=bg_color)

            f = (no - 1) % 12 + 1
            cell = ws_sro.cell(row=row_no, column=start_col + 4)
            cell.value = f
            cell.alignment = CENTER_ALIGN
            cell.font = Font(color=get_fiber_font_color(f - 1))
            bg_color = get_fiber_bg_color(f - 1)
            if bg_color:
                cell.fill = PatternFill(fill_type="solid", start_color=bg_color)

            cell = ws_sro.cell(row=row_no, column=start_col + 5)
            cell.value = f"{extremity_box[BOX_CODE_FIELD_NAME]}-{'%02d' % (i+1)}"
            cell.alignment = CENTER_ALIGN
            if not i:
                cell.hyperlink = Hyperlink(ref=f'{excel_utils.num_to_col(start_col+5)}{row_no}', location=f"#'{extremity_box[BOX_CODE_FIELD_NAME]}'!A1")
                cell.font = Font(name=FONT_NAME, color="0000FF", size=11, underline="single")

        next_segment = self.cable_service.get_next_segment(segment=segment)
        self.fill_segment_and_next_segment_on_d3(ws_sro, row_no + 1, start_col, next_segment)


    def create_return_topo_cell(self, ws_sro, sheet_topo_title):
        cell = ws_sro.cell(row=1, column=1)
        cell.value = f"Return to\n{sheet_topo_title}"
        # return_link = Hyperlink(
        #     ref="A1",  # 当前单元格位置（要设置超链接的单元格）
        #     target=f"#'{sheet_topo_title}'!A1",  # 目标地址：#sheet名!单元格
        #     location=f"#'{sheet_topo_title}'!A1",  # 目标地址：#sheet名!单元格
        #     display=f"Click to jump to {sheet_topo_title}"  # 单元格显示的文本
        # )
        cell.hyperlink = Hyperlink(ref='A1',location=f"#'{sheet_topo_title}'!A1")  # 直接链接到拓扑总览
        # cell.hyperlink = return_link
        cell.font = Font(name=FONT_NAME, color="0000FF", size=11, underline="single")
        cell.alignment = TITLE_WRAP_ALIGN
        ws_sro.column_dimensions[excel_utils.num_to_col(1)].width = 18

    def draw_box_sheet_title(self, ws_box, title, start_col):
        # 2. 基础配置参数
        title_1st_row = 2  # 表头第一行（SRO、Splice State等所在行）
        # start_col = 1  # 起始列（A列，对应数字1）
        # font_name = "宋体"  # 全局字体

        self.create_merged_cell(
            ws=ws_box,
            start_row=1,
            start_col=start_col + 1,
            end_row=1,
            end_col=start_col + 30,
            value=title,
            font=Font(name=FONT_NAME, size=32),
            fill=None,
            alignment=TITLE_BASE_ALIGN
        )


    def draw_box_sheet_header(self, ws_box, start_col):
        # 2. 基础配置参数
        title_1st_row = 2  # 表头第一行（SRO、Splice State等所在行）
        # start_col = 1  # 起始列（A列，对应数字1）
        # font_name = "宋体"  # 全局字体

        for group_idx in range(1, 4):
            self.draw_title_group(ws_box, group_idx, start_col, title_1st_row)
            start_col += COL_LOOP

        self.draw_rest_title(ws_box, start_col, title_1st_row)

        # 调整行高（第2行和第3行，适配合并单元格内容）
        ws_box.row_dimensions[1].height = 45
        ws_box.row_dimensions[title_1st_row].height = 18  # 表头第一行（SRO等）
        ws_box.row_dimensions[title_1st_row + 1].height = 18  # 表头第二行（SRO Port等）
        return title_1st_row + 2


    def draw_rest_title(self, ws, start_col, title_1st_row):
        # _PBO
        cell = ws.cell(row=title_1st_row, column=start_col)
        cell.value = "PBO"
        cell.font = Font(name=FONT_NAME, size=12, bold=True)
        cell.fill = PatternFill(fill_type="solid", start_color=TITLE_BOX_BG_COLOR)
        cell.border = CELL_BORDER
        cell.alignment = TITLE_BASE_ALIGN

        # _PBO_PORT
        cell = ws.cell(row=title_1st_row + 1, column=start_col)
        cell.value = "Port"
        cell.font = Font(name=FONT_NAME, size=11, bold=True)
        cell.fill = PatternFill(fill_type="solid", start_color=TITLE_BOX_BG_COLOR)
        cell.border = CELL_BORDER
        cell.alignment = TITLE_BASE_ALIGN

        # _S
        cell = ws.cell(row=title_1st_row, column=start_col + 1)
        cell.value = "S:"
        cell.font = Font(name=FONT_NAME, size=11, bold=True)
        cell.border = CELL_BORDER
        cell.alignment = TITLE_BASE_ALIGN

        # _R
        cell = ws.cell(row=title_1st_row + 1, column=start_col + 1)
        cell.value = "R:"
        cell.font = Font(name=FONT_NAME, size=11, bold=True)
        cell.border = CELL_BORDER
        cell.alignment = TITLE_BASE_ALIGN

        # _Splice
        cell = ws.cell(row=title_1st_row, column=start_col + 2)
        cell.value = "Splice"
        cell.font = Font(name=FONT_NAME, size=11)
        cell.border = CELL_BORDER
        cell.alignment = TITLE_BASE_ALIGN

        # _Reserve
        cell = ws.cell(row=title_1st_row + 1, column=start_col + 2)
        cell.value = "Reserve"
        cell.font = Font(name=FONT_NAME, size=11)
        cell.border = CELL_BORDER
        cell.alignment = TITLE_BASE_ALIGN

        ws.column_dimensions[excel_utils.num_to_col(start_col)].width = 32


    def draw_title_group(self, ws, group_idx, start_col, title_1st_row):
        section_class = self.gen_section_class(group_idx)
        # 3. 第一组表头（A2:C2 + A3:C3）
        # ----------------------
        # 3.1 合并A2:C2（同一行，列：start_col ~ start_col+2）
        self.create_merged_cell(
            ws=ws,
            start_row=title_1st_row,
            start_col=start_col,
            end_row=title_1st_row,
            end_col=start_col + 2,
            value="SRO",
            font=Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF"),
            fill=PatternFill(fill_type="solid", start_color=TITLE_BOX_BG_COLOR),
            alignment=TITLE_BASE_ALIGN
        )
        # 设置A2单元格样式（合并后以左上角单元格为准）
        cell_a2 = ws.cell(row=title_1st_row, column=start_col)
        cell_a2.value = "SRO" if group_idx == 1 else "BJO"
        cell_a2.font = Font(name=FONT_NAME, size=12, bold=True)  # 白色文字对比绿色背景
        cell_a2.fill = PatternFill(fill_type="solid", start_color=TITLE_BOX_BG_COLOR)
        cell_a2.border = CELL_BORDER
        cell_a2.alignment = TITLE_BASE_ALIGN
        # 3.2 第一组第二行（A3:C3）
        a3_c3_values = ["SRO Port", "ODF Code", "ODF Port"] if group_idx == 1 else ["Code", "Type", "NO."]  # 对应A3、B3、C3的值
        for col_offset in range(3):  # 循环处理A3（offset=0）、B3（offset=1）、C3（offset=2）
            cell = ws.cell(
                row=title_1st_row + 1,
                column=start_col + col_offset
            )
            cell.value = a3_c3_values[col_offset]
            cell.font = Font(name=FONT_NAME, size=11, bold=True)
            cell.fill = PatternFill(fill_type="solid", start_color=TITLE_BOX_BG_COLOR)
            cell.border = CELL_BORDER
            cell.alignment = TITLE_BASE_ALIGN
        # ----------------------
        # 4. 第二组表头（D2:D3 + E2 + E3）- 含自动换行
        # ----------------------
        # 4.1 合并D2:D3（同一列，行：title_1st_row ~ title_1st_row+1）
        self.create_merged_cell(
            ws=ws,
            start_row=title_1st_row,
            start_col=start_col + 3,
            end_row=title_1st_row + 1,
            end_col=start_col + 3,
            value="Splice State",
            font=Font(name=FONT_NAME, size=12, bold=True),
            fill=PatternFill(fill_type="solid", start_color=TITLE_SPLICE_BG_COLOR),
            alignment=TITLE_WRAP_ALIGN
        )
        # 设置D2单元格样式（合并后以左上角单元格为准，启用自动换行）
        cell_d2 = ws.cell(row=title_1st_row, column=start_col + 3)
        cell_d2.value = "Splice State"  # 内容过长时会自动换行（也可手动加\n强制换行）
        cell_d2.font = Font(name=FONT_NAME, size=12, bold=True)
        cell_d2.fill = PatternFill(fill_type="solid", start_color=TITLE_SPLICE_BG_COLOR)
        cell_d2.border = CELL_BORDER
        cell_d2.alignment = TITLE_WRAP_ALIGN  # 关键：应用自动换行对齐
        # 4.2 处理E2和E3（空白单元格，仅保留样式）
        for row_offset in [0, 1]:  # 循环处理E2（row_offset=0）、E3（row_offset=1）
            cell = ws.cell(
                row=title_1st_row + row_offset,
                column=start_col + 4
            )
            cell.fill = PatternFill(fill_type="solid", start_color=TITLE_BLANK_BG_COLOR)
            cell.border = CELL_BORDER
            cell.alignment = TITLE_BASE_ALIGN
        # ----------------------
        # 5. 第三组表头（F2:J2 + F3:J3）
        # ----------------------
        # 5.1 合并F2:J2（同一行，列：start_col+5 ~ start_col+9）
        self.create_merged_cell(
            ws=ws,
            start_row=title_1st_row,
            start_col=start_col + 5,
            end_row=title_1st_row,
            end_col=start_col + 9,
            value="Distribution 1",
            font=Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF"),
            fill=PatternFill(fill_type="solid", start_color=TITLE_SECTION_BG_COLOR),
            alignment=TITLE_BASE_ALIGN
        )
        # 设置F2单元格样式
        cell_f2 = ws.cell(row=title_1st_row, column=start_col + 5)
        cell_f2.value = section_class
        cell_f2.font = Font(name=FONT_NAME, size=12, bold=True)  # 白色文字对比蓝色背景
        cell_f2.fill = PatternFill(fill_type="solid", start_color=TITLE_SECTION_BG_COLOR)
        cell_f2.border = CELL_BORDER
        cell_f2.alignment = TITLE_BASE_ALIGN
        # 5.2 第三组第二行（F3:J3）
        f3_j3_values = ["Section", "Type", "NO.", "T", "F"]  # 对应F3、G3、H3、I3、J3的值
        for col_offset in range(5):  # 循环处理5列
            cell = ws.cell(
                row=title_1st_row + 1,
                column=start_col + 5 + col_offset
            )
            cell.value = f3_j3_values[col_offset]
            cell.font = Font(name=FONT_NAME, size=11, bold=True)
            cell.fill = PatternFill(fill_type="solid", start_color=TITLE_SECTION_BG_COLOR)
            cell.border = CELL_BORDER
            cell.alignment = TITLE_BASE_ALIGN
        # ----------------------
        # 6. 优化列宽和行高（避免内容溢出）
        # ----------------------
        # 调整列宽（A-J列，根据内容长度适配）

        col_widths = [12 if group_idx == 1 else 32,
                      12 if group_idx == 1 else 10,
                      12 if group_idx == 1 else 10,
                      10, 4, 8, 8, 6, 6, 6
                      ]
        for col_idx in range(0, len(col_widths)):  # 1=A列，10=J列
            ws.column_dimensions[excel_utils.num_to_col(col_idx + start_col)].width = col_widths[col_idx]


    def gen_section_class(self, _class: int):
        return f"Distribution {_class}"


    def create_merged_cell(self, ws, start_row, start_col, end_row, end_col, value, font, alignment, fill=None):
        """创建合并单元格并应用样式和边框"""
        # 合并单元格
        ws.merge_cells(start_row=start_row, start_column=start_col,
                       end_row=end_row, end_column=end_col)
        # 设置内容和样式（仅需设置左上角单元格）
        cell = ws.cell(row=start_row, column=start_col)
        cell.value = value
        cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = alignment
        # 设置四周边框（关键：遍历所有单元格设置边缘）
        self.set_merged_border(ws, start_row, start_col, end_row, end_col)


    def set_merged_border(self, ws, start_row, start_col, end_row, end_col):
        """为合并区域设置完整四周边框"""
        # 遍历合并区域的所有单元格
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                # 初始化边框
                border = Border()

                # 顶部边框：只给第一行设置上边框
                if row == start_row:
                    border.top = BORDER_SIDE
                # 底部边框：只给最后一行设置下边框
                if row == end_row:
                    border.bottom = BORDER_SIDE
                # 左侧边框：只给第一列设置左边框
                if col == start_col:
                    border.left = BORDER_SIDE
                # 右侧边框：只给最后一列设置右边框
                if col == end_col:
                    border.right = BORDER_SIDE

                cell.border = border


COL_LOOP = 10
ODF_MAX_PORT_NO = 144
TARGET_ROW = 480
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
# 对齐方式配置（水平居中+垂直居中，第二组额外加自动换行）
TITLE_BASE_ALIGN = Alignment(horizontal="center", vertical="center")  # 基础对齐（无自动换行）
TITLE_WRAP_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 带自动换行的对齐

BOX_SHEET_TABLE_START_COL = 1
# 边框样式（统一黑色细边框）
BORDER_SIDE = Side(style="thin", color="000000")
CELL_BORDER = Border(
    left=BORDER_SIDE,
    right=BORDER_SIDE,
    top=BORDER_SIDE,
    bottom=BORDER_SIDE
)
SHEET_ZOOM_SCALE = 70
# 颜色定义（RGB格式）
TITLE_BOX_BG_COLOR = "00B050"  # 第一组绿色背景
TITLE_SPLICE_BG_COLOR = "FFFF00"  # 第二组黄色背景
TITLE_BLANK_BG_COLOR = "FFFFFF"  # 第三组蓝色背景
TITLE_SECTION_BG_COLOR = "00B0F0"  # 第三组蓝色背景

FIBER_COLOR = {
    0: {"bgc": "FF0000", "fc": "FFFFFF"},  # idx=1 红
    1: {"bgc": "0066FF", "fc": "FFFFFF"},  # idx=2 蓝
    2: {"bgc": "00B050", "fc": "FFFFFF"},  # idx=3 绿
    3: {"bgc": "FFFF00", "fc": "000000"},  # idx=4 黄
    4: {"bgc": "7030A0", "fc": "FFFFFF"},  # idx=5 紫
    5: {"bgc": None, "fc": "000000"},  # idx=6 白
    6: {"bgc": "FFCC00", "fc": "000000"},  # idx=7 金
    7: {"bgc": "D9D9D9", "fc": "000000"},  # idx=8 灰
    8: {"bgc": "963634", "fc": "FFFFFF"},  # idx=9 赭
    9: {"bgc": "FDE9D9", "fc": "000000"},  # idx=10 米
    10: {"bgc": "66FFFF", "fc": "000000"},  # idx=11 青
    11: {"bgc": "FFCCFF", "fc": "000000"}  # idx=12 粉
}


def get_fiber_bg_color(no: int):
    return FIBER_COLOR[no % 12]['bgc']


def get_fiber_font_color(no: int):
    return FIBER_COLOR[no % 12]['fc']
