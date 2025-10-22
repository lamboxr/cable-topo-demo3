import openpyxl
import pandas as pd
from openpyxl.styles import Border, Side, Font, Alignment

import init_data
from constraints.field_name_mapper import *
from data_service import data_service_box, data_service_cable, data_service_sro
from data_service.data_service_cable import get_next_segment_by_origin_code
from utils import excel_utils

# 全局配置
COLUMN_WIDTHS = {
    'A': 32,
    'B': 60,
    'C': 48,
    'D': 60,
    'E': 48,
    'F': 60,
    'G': 48
}
LEVEL_TO_COLUMN = {
    1: 'B',
    2: 'D',
    3: 'F'
}
GROUP_ROWS = 10  # 每组占用8行
SHEET_ZOOM_SCALE = 50


def init_workbook():
    """初始化Excel工作簿，设置列宽和样式基础"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "逻辑拓扑图"
    ws.sheet_view.zoomScale = SHEET_ZOOM_SCALE

    # 设置列宽
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    return wb, ws


def generate_topo_excel(output_path):
    """生成拓扑Excel的主函数"""
    wb, ws = init_workbook()
    current_row = 2  # 起始行

    # 1. 获取所有SRO节点（根节点），按code升序
    sro_boxes = data_service_sro.get_all_sro_order_by_code_asc()
    if sro_boxes is None or sro_boxes.empty:
        print("无SRO节点数据，无法生成拓扑图")
        return

    for _, sro in sro_boxes.iterrows():
        current_row = draw_point_and_resources(ws=ws, current_row=current_row, box_data=sro, upper_cable_level=0,
                                               upper_section_value="N/A")
        sub_cable_amt = data_service_cable.get_all_first_segments_start_with_box_order_by_code_asc(
            sro[BOX_CODE_FIELD_NAME], "N/A")
        if sub_cable_amt.empty:
            current_row += GROUP_ROWS

    # 保存文件
    wb.save(output_path)
    print(f"拓扑图已生成：{output_path}")


def draw_point_and_resources(ws, current_row, box_data, upper_cable_level, upper_section_value):
    """判断点是否需要描绘分支线，至少两个子线缆时，初始化点的分支线的描绘开关、起止点行数"""
    (need_to_draw_box_vertical_branch_line,
     box_vertical_branch_line_start_row) = does_need_to_draw_box_vertical_branch_line(
        current_row, box_data)

    """==================BOX描绘开始=================="""
    """描绘BOX点"""
    current_row = draw_box_node_and_next_segment_in_same_section(ws, current_row, box_data, upper_cable_level,
                                                                 upper_section_value,
                                                                 need_to_draw_box_vertical_branch_line)

    """查询以该BOX为起点, 但section!=box的section的所有第一段segments,"""
    first_segments_list = data_service_cable.get_all_first_segments_start_with_box_order_by_code_asc(
        box_data[BOX_CODE_FIELD_NAME], upper_section_value)

    # """查询BOX上的子线缆列表"""
    # first_section_list = data_service_cable.get_all_cables_start_with_one_point_order_by_code_asc(box_data['code'])

    if first_segments_list is None or first_segments_list.empty:
        """如果当前点没有子线缆，行号直接下移一个描绘空间
        用于描绘同级下一个点的行定位
        或者描绘循环中最后一个点的上一级线缆的同级下一个线缆的行定位"""
        return current_row + GROUP_ROWS
    else:
        """==================子线缆描绘开始=================="""
        sub_cables_amt = len(first_segments_list)
        sub_cable_idx = 0
        for _, _1st_segment in first_segments_list.iterrows():
            sub_cable_idx += 1
            if need_to_draw_box_vertical_branch_line and sub_cable_idx == sub_cables_amt:
                """当需要画竖向分支线，且即将开始画最后一条线缆前：画竖向分支线"""
                draw_box_vertical_branch_line(ws=ws, upper_cable_level=upper_cable_level,
                                              start_row=box_vertical_branch_line_start_row,
                                              end_row=current_row + 1)
            """判断点是否需要描绘线缆的竖向路径，至少两个pass点时，初始化点的线缆竖向线条的描绘开关，与起止点行数"""
            (need_to_draw_cable_vertical_route_line,
             cable_vertical_route_start_row) = does_need_to_draw_cable_vertical_right_line(current_row, _1st_segment)
            """开始描绘线缆第一个section横向部分(和自身描绘空间的右侧边框 bool参数)"""
            current_row = draw_1st_segment(ws, current_row, _1st_segment, upper_cable_level,
                                           need_to_draw_cable_vertical_route_line)

            """==================子线缆描绘完成=================="""

            """==================子线缆掏芯点开始=================="""
            """查询线缆上的掏芯点列表"""
            sub_boxes_on_complete_cable = get_sub_boxes_on_one_complete_cable_start_with_1st_section(_1st_segment, [])

            sub_boxes_amt = len(sub_boxes_on_complete_cable)
            if sub_boxes_on_complete_cable is None or not sub_boxes_on_complete_cable:
                """线缆上无掏心点，进入下一个线缆，下移一个描绘空间画同级的next线缆"""
                current_row += GROUP_ROWS
                continue
            elif sub_boxes_amt >= 2:
                # if need_to_draw_cable_vertical_route_line:
                """线缆上至少2个掏心点，下移一个描绘空间开始画点"""
                current_row += GROUP_ROWS

            sub_box_idx = 0
            for box_on_cable in sub_boxes_on_complete_cable:
                sub_box_idx += 1
                if need_to_draw_cable_vertical_route_line and sub_box_idx == sub_boxes_amt:
                    """当需要画竖向路由线，即将画最后一个掏芯点前：开始画竖向路由线"""
                    draw_cable_vertical_route_line(ws=ws, cable=_1st_segment, start_row=cable_vertical_route_start_row,
                                                   end_row=current_row)
                """||||||||||继续调用递归函数描绘点||||||||||"""
                current_row = draw_point_and_resources(ws=ws, current_row=current_row, box_data=box_on_cable,
                                                       upper_cable_level=_1st_segment[CABLE_LEVEL_FIELD_NAME],
                                                       upper_section_value=_1st_segment[CABLE_SECTION_FIELD_NAME])

            """==================子线缆掏芯点结束=================="""
    return current_row


def get_sub_boxes_on_one_complete_cable_start_with_1st_section(segment, sub_boxes_list: list):
    if segment is not None and not segment.empty:
        section = segment[CABLE_SECTION_FIELD_NAME]
        box = data_service_box.get_box_by_code(segment[CABLE_EXTREMITY_FIELD_NAME])
        if not box.empty:
            sub_boxes_list.append(box)
            next_segment = data_service_cable.get_next_segment_by_origin_code(section, box[BOX_CODE_FIELD_NAME])
            return get_sub_boxes_on_one_complete_cable_start_with_1st_section(next_segment, sub_boxes_list)
        else:
            return sub_boxes_list
    else:
        return sub_boxes_list


def draw_box_node_and_next_segment_in_same_section(ws, start_row, box_data, upper_cable_level,
                                                   upper_section, need_to_draw_box_vertical_branch_line):
    col = "A" if upper_cable_level == 0 else excel_utils.get_right_col_letter(
        LEVEL_TO_COLUMN[upper_cable_level])
    sheet_name = create_box_sheet(ws, col, box_data)
    # 子sheet内容暂时留空（无需额外操作）
    current_row = draw_box_node(ws, start_row, col, box_data, sheet_name, need_to_draw_box_vertical_branch_line)
    next_segment = get_next_segment_by_origin_code(upper_section, box_data[BOX_CODE_FIELD_NAME])
    if next_segment is not None and not next_segment.empty:
        draw_vertical_segment(ws, current_row, col, next_segment)
    return current_row


def create_box_sheet(ws, col, box_data):
    if col in ['A', 'G']:
        sheet_name = box_data[BOX_CODE_FIELD_NAME]
        # 获取工作簿对象（从当前工作表反向获取）
        wb = ws.parent
        # 检查子sheet是否已存在，不存在则创建
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append([])
            overview_cell = ws.cell(row=1, column=1)
            overview_cell.value = "返回逻辑拓扑图"
            overview_cell.hyperlink = "#'逻辑拓扑图'!A1"  # 直接链接到拓扑总览
            overview_cell.font = Font(color="0000FF", underline="single")
            return sheet_name
    return None


def draw_box_node(ws, start_row, col, box_data, sheet_name, need_to_draw_box_vertical_branch_line):
    # if box_data['class'] == 'SRO':
    #     return draw_sro_node_(ws, start_row, box_data, need_to_draw_box_vertical_branch_line)
    # else:
    #     return draw_closure_pbo_node_(ws, start_row, box_data, upper_cable_level, need_to_draw_box_vertical_branch_line)
    """确定下一级节点的列（线缆列的右侧列：0-A, B→C，D→E，F→G）"""

    """绘制C/E/G列的Closure/PBO节点（8行一组，不合并单元格）"""
    # 第1-4行：粗外侧线框（模拟盒子）
    for row in range(start_row, start_row + 4):
        # 第1行：nap.class（加粗居中）
        if row == start_row:
            set_cell(
                ws,
                row=row,
                col=col,
                value=f"TYPE:{box_data[BOX_CLASS_FIELD_NAME]}",
                border=BOX_1ST_ROW_BORDER,
                font=BOLD_FONT,
                align=CENTER_ALIGN
            )

        # 第2行：nap.code + "    " + nap.type（居中）
        elif row == start_row + 1:
            set_cell(
                ws,
                row=row,
                col=col,
                value=f"{box_data[BOX_CODE_FIELD_NAME]}",
                border=BOX_MIDDLE_ROW_BORDER,
                align=CENTER_ALIGN
            )
            print(f"{row} , {col} , {box_data[BOX_CLASS_FIELD_NAME]}")
            if sheet_name:
                cell = ws.cell(row=row, column=excel_utils.col_to_num(col))
                cell.hyperlink = f"#'{sheet_name}'!A1"
                cell.font = Font(color="0000FF", underline="single")
        # 第3行：留空
        elif row == start_row + 2:
            set_cell(
                ws,
                row=row,
                col=col,
                value=f"{box_data[BOX_TYPE_FIELD_NAME]}" if col != "A" else None,
                border=BOX_MIDDLE_ROW_BORDER,
                align=CENTER_ALIGN
            )
        # 第4行：nap.in_start + "-" + nap.in_end（居中）
        elif row == start_row + 3:
            set_cell(
                ws,
                row=row,
                col=col,
                value=None,
                border=BOX_LAST_ROW_BORDER,
                align=CENTER_ALIGN
            )
    # 第5-8行：留空（无边框）
    for row in range(start_row + 4, start_row + GROUP_ROWS):
        set_cell(ws, row=row, col=col, value=None,
                 border=CABLE_ROUTE_BORDER if need_to_draw_box_vertical_branch_line else Border())
    return start_row  # 移动到下一组


def draw_sro_node(ws, start_row, sro_data, need_to_draw_box_vertical_branch_line):
    """绘制A列的SRO节点（8行一组，不合并单元格）"""
    # 第1-4行：粗外侧线框（模拟盒子）
    for row in range(start_row, start_row + 4):  # 行索引：start_row到start_row+3
        # 第1行：nap.class（加粗居中）
        if row == start_row:
            set_cell(
                ws,
                row=row,
                col='A',
                value=sro_data['class'],
                border=BOX_1ST_ROW_BORDER,
                font=BOLD_FONT,
                align=CENTER_ALIGN
            )
        # 第2行：nap.code（居中）
        elif row == start_row + 1:
            set_cell(
                ws,
                row=row,
                col='A',
                value=sro_data['code'],
                border=BOX_MIDDLE_ROW_BORDER,
                align=CENTER_ALIGN
            )
        elif row == start_row + 2:
            set_cell(
                ws,
                row=row,
                col='A',
                value=None,
                border=BOX_MIDDLE_ROW_BORDER
            )
        # 第3-4行：留空（保持边框）
        else:
            set_cell(
                ws,
                row=row,
                col='A',
                value=None,
                border=BOX_LAST_ROW_BORDER
            )
    # 第5-8行：留空（无边框）
    for row in range(start_row + 4, start_row + GROUP_ROWS):
        set_cell(ws, row=row, col='A', value=None,
                 border=CABLE_ROUTE_BORDER if need_to_draw_box_vertical_branch_line else Border())
    return start_row  # 移动到下一组


def draw_vertical_segment(ws, start_row, col, segment):
    section = segment[CABLE_SECTION_FIELD_NAME]
    code = segment[CABLE_CODE_FIELD_NAME]
    type = segment[CABLE_TYPE_FIELD_NAME]
    length = segment[CABLE_LENGTH_FIELD_NAME]
    line_5_right_border = get_cell_right_boarder(ws, start_row + 5, col)
    if line_5_right_border:
        border = Border(
            right=Side(style=THICKER_WIDTH),
            left=Side(style=THICKER_WIDTH)
        )
    else:
        border = Border()

    set_cell(ws, row=start_row + 6, col=col, value=f"{section} / {code}",
             border=border, align=LEFT_ALIGN, font=BOLD_FONT)
    set_cell(ws, row=start_row + 7, col=col, value=f"{type} {length}",
             border=border, align=LEFT_ALIGN)


def draw_closure_pbo_node(ws, start_row, box_data, upper_cable_level, need_to_draw_box_vertical_branch_line):
    # 3. 确定下一级节点的列（线缆列的右侧列：B→C，D→E，F→G）
    current_box_col = excel_utils.get_right_col_letter(LEVEL_TO_COLUMN[upper_cable_level])
    """绘制C/E/G列的Closure/PBO节点（8行一组，不合并单元格）"""
    # 第1-4行：粗外侧线框（模拟盒子）
    for row in range(start_row, start_row + 4):
        # 第1行：nap.class（加粗居中）
        if row == start_row:
            set_cell(
                ws,
                row=row,
                col=current_box_col,
                value=box_data['class'],
                border=BOX_1ST_ROW_BORDER,
                font=BOLD_FONT,
                align=CENTER_ALIGN
            )
        # 第2行：nap.code + "    " + nap.type（居中）
        elif row == start_row + 1:
            set_cell(
                ws,
                row=row,
                col=current_box_col,
                value=f"{box_data['code']}    {box_data['type']}",
                border=BOX_MIDDLE_ROW_BORDER,
                align=CENTER_ALIGN
            )
        # 第3行：留空
        elif row == start_row + 2:
            set_cell(
                ws,
                row=row,
                col=current_box_col,
                value=f"On:{box_data['cable_in']}",
                border=BOX_MIDDLE_ROW_BORDER,
                align=CENTER_ALIGN
            )
        # 第4行：nap.in_start + "-" + nap.in_end（居中）
        elif row == start_row + 3:
            set_cell(
                ws,
                row=row,
                col=current_box_col,
                value=f"InRange:{int(box_data['in_start'])}-{int(box_data['in_end'])}",
                border=BOX_LAST_ROW_BORDER,
                align=CENTER_ALIGN
            )
    # 第5-8行：留空（无边框）
    for row in range(start_row + 4, start_row + GROUP_ROWS):
        set_cell(ws, row=row, col=current_box_col, value=None,
                 border=CABLE_ROUTE_BORDER if need_to_draw_box_vertical_branch_line else Border())
    return start_row  # 移动到下一组


def draw_cable(ws, start_row, cable_data, upper_cable_level, need_to_draw_cable_vertical_right_line):
    """绘制B/D/F列的线缆（8行一组，不合并单元格）"""
    # 根据level确定列（B=1, D=2, F=3）
    current_level = cable_data['level']
    col = LEVEL_TO_COLUMN[current_level]
    # level显示文本映射
    level_text = {1: "Distribution 01", 2: "Distribution 02", 3: "Distribution 03"}[current_level]

    if current_level - upper_cable_level > 1:
        set_cell(
            ws,
            row=start_row,
            col=excel_utils.get_left_col_letter(col),
            value="",
            border=CABLE_FIRST_ROW_BORDER
        )
        set_cell(
            ws,
            row=start_row,
            col=excel_utils.get_left_col_letter(col, 2),
            value="",
            border=CABLE_FIRST_ROW_BORDER
        )
    route_border = Border()
    if need_to_draw_cable_vertical_right_line:
        route_border = CABLE_ROUTE_BORDER

    # 第1行：level文本（下边框加粗）
    set_cell(
        ws,
        row=start_row,
        col=col,
        value=level_text,
        border=CABLE_FIRST_ROW_BORDER,
        font=BOLD_FONT,
        align=CENTER_ALIGN
    )
    # 第2行：cable.code + "    " + cable.type
    set_cell(
        ws,
        row=start_row + 1,
        col=col,
        value=f"{cable_data['code']}    {cable_data['type']}",
        align=CENTER_ALIGN,
        border=route_border
    )
    # 第3行：cable.r_nodes
    set_cell(
        ws,
        row=start_row + 2,
        col=col,
        value=f"From: {cable_data['origin_box']}    RNodes:{cable_data['r_nodes']}",
        align=CENTER_ALIGN,
        border=route_border
    )
    # 第4行：cable.port_start + "-" + cable.port_end
    set_cell(
        ws,
        row=start_row + 3,
        col=col,
        value=f"PortRange:{int(cable_data['port_start'])}-{int(cable_data['port_end'])}",
        align=CENTER_ALIGN,
        border=route_border
    )
    # 第5-8行：留空
    for row in range(start_row + 4, start_row + GROUP_ROWS):
        set_cell(ws, row=row, col=col, value=None, border=route_border)
    return start_row  # 移动到下一组


def draw_1st_segment(ws, start_row, section_data, upper_cable_level, need_to_draw_cable_vertical_right_line) -> int:
    current_level = section_data[CABLE_LEVEL_FIELD_NAME]
    level_text = {1: "Distribution 01", 2: "Distribution 02", 3: "Distribution 03"}[current_level]

    col = LEVEL_TO_COLUMN[current_level]
    section = section_data[CABLE_SECTION_FIELD_NAME]
    code = section_data[CABLE_CODE_FIELD_NAME]
    type = section_data[CABLE_TYPE_FIELD_NAME]
    length = section_data[CABLE_LENGTH_FIELD_NAME]
    if current_level - upper_cable_level > 1:
        set_cell(
            ws,
            row=start_row,
            col=excel_utils.get_left_col_letter(col),
            value="",
            border=CABLE_FIRST_ROW_BORDER
        )
        set_cell(
            ws,
            row=start_row,
            col=excel_utils.get_left_col_letter(col, 2),
            value="",
            border=CABLE_FIRST_ROW_BORDER
        )
    route_border = Border()
    if need_to_draw_cable_vertical_right_line:
        route_border = CABLE_ROUTE_BORDER

    # 第1行：level文本（下边框加粗）
    set_cell(
        ws,
        row=start_row,
        col=col,
        value=f"{section} / {code}",
        border=CABLE_FIRST_ROW_BORDER,
        font=BOLD_FONT,
        align=CENTER_ALIGN
    )
    # 第2行：cable.code + "    " + cable.type
    set_cell(
        ws,
        row=start_row + 1,
        col=col,
        value=f"{type}    {length}",
        align=CENTER_ALIGN,
        border=route_border
    )
    # # 第3行：cable.r_nodes
    # set_cell(
    #     ws,
    #     row=start_row + 2,
    #     col=col,
    #     value=f"From: {section_data['origin_box']}    RNodes:{section_data['r_nodes']}",
    #     align=CENTER_ALIGN,
    #     border=route_border
    # )
    # # 第4行：cable.port_start + "-" + cable.port_end
    # set_cell(
    #     ws,
    #     row=start_row + 3,
    #     col=col,
    #     value=f"PortRange:{int(section_data['port_start'])}-{int(section_data['port_end'])}",
    #     align=CENTER_ALIGN,
    #     border=route_border
    # )
    # 第3-8行：留空
    for row in range(start_row + 2, start_row + GROUP_ROWS):
        set_cell(ws, row=row, col=col, value=None, border=route_border)
    return start_row  # 移动到下一组


def draw_section(ws, start_row, section_data, is_first_section, upper_cable_level,
                 need_to_draw_cable_vertical_right_line):
    """绘制B/D/F列的线缆（8行一组，不合并单元格）"""
    # 根据level确定列（B=1, D=2, F=3）
    current_level = section_data['level']
    col = LEVEL_TO_COLUMN[current_level]
    # level显示文本映射
    level_text = {1: "Distribution 01", 2: "Distribution 02", 3: "Distribution 03"}[current_level]

    if current_level - upper_cable_level > 1:
        set_cell(
            ws,
            row=start_row,
            col=excel_utils.get_left_col_letter(col),
            value="",
            border=CABLE_FIRST_ROW_BORDER
        )
        set_cell(
            ws,
            row=start_row,
            col=excel_utils.get_left_col_letter(col, 2),
            value="",
            border=CABLE_FIRST_ROW_BORDER
        )
    route_border = Border()
    if need_to_draw_cable_vertical_right_line:
        route_border = CABLE_ROUTE_BORDER

    # 第1行：level文本（下边框加粗）
    set_cell(
        ws,
        row=start_row,
        col=col,
        value=level_text,
        border=CABLE_FIRST_ROW_BORDER,
        font=BOLD_FONT,
        align=CENTER_ALIGN
    )
    # 第2行：cable.code + "    " + cable.type
    set_cell(
        ws,
        row=start_row + 1,
        col=col,
        value=f"{section_data['code']}    {section_data['type']}",
        align=CENTER_ALIGN,
        border=route_border
    )
    # 第3行：cable.r_nodes
    set_cell(
        ws,
        row=start_row + 2,
        col=col,
        value=f"From: {section_data['origin_box']}    RNodes:{section_data['r_nodes']}",
        align=CENTER_ALIGN,
        border=route_border
    )
    # 第4行：cable.port_start + "-" + cable.port_end
    set_cell(
        ws,
        row=start_row + 3,
        col=col,
        value=f"PortRange:{int(section_data['port_start'])}-{int(section_data['port_end'])}",
        align=CENTER_ALIGN,
        border=route_border
    )
    # 第5-8行：留空
    for row in range(start_row + 4, start_row + GROUP_ROWS):
        set_cell(ws, row=row, col=col, value=None, border=route_border)
    return start_row  # 移动到下一组


def draw_box_vertical_branch_line(ws, upper_cable_level, start_row, end_row):
    if upper_cable_level == 0:
        col = 'A'
    else:
        col = excel_utils.get_right_col_letter(LEVEL_TO_COLUMN[upper_cable_level])

    for row in range(start_row, end_row):
        set_cell(
            ws,
            row=row,
            col=col,
            value="",
            border=CABLE_ROUTE_BORDER
        )


def draw_cable_vertical_route_line(ws, cable, start_row, end_row):
    col = LEVEL_TO_COLUMN[cable[CABLE_LEVEL_FIELD_NAME]]
    for row in range(start_row, end_row):
        set_cell(
            ws,
            row=row,
            col=col,
            value="",
            border=CABLE_ROUTE_BORDER
        )


def does_need_to_draw_box_vertical_branch_line(current_row, box_data):
    sub_cables_amt = data_service_cable.get_sub_cables_amt(box_data[BOX_CODE_FIELD_NAME])
    if sub_cables_amt >= 2:
        return True, current_row + GROUP_ROWS
    else:
        return False, None


def does_need_to_draw_cable_vertical_right_line(current_row, _1st_segment_data):
    has_at_least_2_sections = data_service_cable.has_at_least_2_segments_on_cable(_1st_segment_data)
    if has_at_least_2_sections:
        return True, current_row + GROUP_ROWS
    else:
        return False, None


THIN_WIDTH = 'thin'
THICKER_WIDTH = 'medium'

# 样式定义调整：为独立单元格设置完整边框
BOX_1ST_ROW_BORDER = Border(
    left=Side(style=THICKER_WIDTH),
    right=Side(style=THICKER_WIDTH),
    top=Side(style=THICKER_WIDTH),
    bottom=Side(style=THIN_WIDTH)
)

BOX_MIDDLE_ROW_BORDER = Border(
    left=Side(style=THICKER_WIDTH),
    right=Side(style=THICKER_WIDTH),
)

BOX_LAST_ROW_BORDER = Border(
    left=Side(style=THICKER_WIDTH),
    right=Side(style=THICKER_WIDTH),
    bottom=Side(style=THICKER_WIDTH)
)

# 线缆第一行下边框加粗（其他边框默认）
CABLE_FIRST_ROW_BORDER = Border(
    bottom=Side(style=THICKER_WIDTH)
)

CABLE_ROUTE_BORDER = Border(
    right=Side(style=THICKER_WIDTH)
)

POINT_BRANCHES_BORDER = Border(
    right=Side(style=THICKER_WIDTH)
)

BOLD_FONT = Font(bold=True)
BOLD_LINK_FONT = Font(bold=True, italic=True, color="0000FF", underline="single")
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)


def merge_cells(ws, start_row, end_row, col, value, border=None, font=None, align=None):
    """
    合并单元格并设置内容和样式
    :param ws: 工作表
    :param start_row: 起始行
    :param end_row: 结束行
    :param col: 列标识（如'A'）
    :param value: 单元格内容
    :param border: 边框样式
    :param font: 字体样式
    :param align: 对齐方式
    """
    cell_range = f"{col}{start_row}:{col}{end_row}"
    # ws.merge_cells(cell_range)
    cell = ws[f"{col}{start_row}"]
    cell.value = value
    cell.border = border or Border()
    cell.font = font or Font()
    cell.alignment = align or Alignment()


def get_cell_right_boarder(ws, row, col):
    cell = ws[f"{col}{row}"]

    # 获取单元格的边框对象
    return cell.border.right


def set_cell(ws, row, col, value, border=None, font=None, align=None):
    """设置单个单元格内容和样式"""
    cell = ws[f"{col}{row}"]
    cell.value = value
    cell.border = border or Border()
    cell.font = font or Font()
    cell.alignment = align or Alignment()


# def get_str_value(value):
#     if value is None:
#         return None
#     elif isinstance(value, pd.Series):
#         # 若Series非空，取第一个值；否则返回空条件
#         if not value.empty:
#             return value.iloc[0]  # 提取第一个元素作为标量
#         return None
#     return value

if __name__ == '__main__':
    generate_topo_excel("拓扑图输出2.xlsx")
